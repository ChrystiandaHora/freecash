"""
Serviço de importação de relatórios Excel (.xlsx).
Permite ler, validar e importar movimentações de um arquivo de relatório exportado.
"""

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from core.models import Conta, Categoria
from investimento.models import Ativo, Transacao as TransacaoInvestimento


def parse_date_br(date_str: str):
    """Parse date in dd/mm/yyyy format."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str.strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def parse_valor(valor):
    """Parse valor from Excel (pode ser float ou string)."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    try:
        # Remove formatação brasileira
        valor_str = str(valor).strip().replace(".", "").replace(",", ".")
        return Decimal(valor_str)
    except (InvalidOperation, ValueError):
        return None


def ler_excel_relatorio(arquivo) -> dict:
    """
    Lê arquivo Excel de relatório e retorna dados estruturados.

    Returns:
        dict com keys:
            - 'dados': lista de dicionários com os registros
            - 'resumo': dict com total_receitas, total_despesas, quantidade
            - 'erro': mensagem de erro se houver problema
    """
    try:
        # Lê o arquivo
        if hasattr(arquivo, "read"):
            content = arquivo.read()
            arquivo.seek(0)  # Reset para uso posterior
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        else:
            wb = load_workbook(filename=io.BytesIO(arquivo), read_only=True)

        ws = wb.active

        # Encontra linha de cabeçalho (procura por "Data" na primeira coluna)
        header_row = None
        for row_num, row in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), 1
        ):
            if row and row[0] and str(row[0]).strip().upper() == "DATA":
                header_row = row_num
                break

        if header_row is None:
            return {"erro": "Formato inválido: cabeçalho 'Data' não encontrado."}

        # Valida colunas esperadas - extrai valores das células
        header_cells = list(
            ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )[0]
        headers = [str(cell).strip() if cell else "" for cell in header_cells]
        expected = ["Data", "Tipo", "Descrição", "Categoria", "Valor (R$)", "Status"]

        # Verifica se as primeiras 6 colunas correspondem (case insensitive)
        for i, exp in enumerate(expected):
            if i >= len(headers):
                return {"erro": f"Formato inválido: coluna '{exp}' não encontrada."}
            if headers[i].upper() != exp.upper():
                # Aceita variações comuns
                if exp == "Valor (R$)" and "VALOR" in headers[i].upper():
                    continue
                return {
                    "erro": f"Formato inválido: esperado '{exp}', encontrado '{headers[i]}'."
                }

        # Lê os dados
        dados = []
        total_receitas = Decimal("0.00")
        total_despesas = Decimal("0.00")

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Ignora linhas vazias ou de resumo
            if not row or not row[0]:
                continue

            # Ignora linhas de resumo (Total Receitas, Total Despesas, Saldo)
            primeira_col = str(row[0]).strip() if row[0] else ""
            if primeira_col.upper() in [
                "TOTAL RECEITAS:",
                "TOTAL DESPESAS:",
                "SALDO:",
                "",
            ]:
                continue
            if any(x in primeira_col.upper() for x in ["TOTAL", "SALDO"]):
                continue

            # Parse data
            data = row[0]
            if isinstance(data, datetime):
                data = data.date()
            elif isinstance(data, str):
                data = parse_date_br(data)

            if not data:
                continue

            tipo_str = str(row[1]).strip() if row[1] else ""
            descricao = str(row[2]).strip() if row[2] else ""
            categoria_nome = str(row[3]).strip() if row[3] else ""
            valor = parse_valor(row[4])
            status_str = str(row[5]).strip() if row[5] else ""

            if not valor or valor <= 0:
                continue

            # Determina tipo
            tipo = (
                Conta.TIPO_RECEITA
                if tipo_str.upper() == "RECEITA"
                else Conta.TIPO_DESPESA
            )

            # Determina status
            realizada = status_str.upper() in [
                "REALIZADA",
                "OK",
                "PAGO",
                "PAGA",
                "SIM",
                "TRUE",
                "1",
            ]

            registro = {
                "data": data,
                "tipo": tipo,
                "tipo_label": tipo_str,
                "descricao": descricao,
                "categoria_nome": categoria_nome,
                "valor": valor,
                "realizada": realizada,
                "status_label": "Realizada" if realizada else "Pendente",
            }
            dados.append(registro)

            if tipo == Conta.TIPO_RECEITA:
                total_receitas += valor
            else:
                total_despesas += valor

        wb.close()

        return {
            "dados": dados,
            "resumo": {
                "total_receitas": total_receitas,
                "total_despesas": total_despesas,
                "saldo": total_receitas - total_despesas,
                "quantidade": len(dados),
            },
            "erro": None,
        }

    except Exception as e:
        return {"erro": f"Erro ao ler arquivo: {str(e)}"}


def validar_dados_relatorio(dados: list, usuario) -> list:
    """
    Valida os dados do relatório contra o banco de dados.
    Retorna lista de dados com informações de validação adicionadas.

    Adiciona aos registros:
        - 'categoria_id': ID da categoria encontrada ou None
        - 'categoria_existe': bool
        - 'existe_duplicado': bool (se já existe conta com mesma data/descrição/valor)
        - 'conta_existente_id': ID da conta existente ou None
    """
    # Busca categorias do usuário
    categorias = {c.nome.upper(): c for c in Categoria.objects.filter(usuario=usuario)}

    # Busca contas existentes para verificar duplicados
    contas_existentes = {
        (c.data_prevista, c.descricao.upper(), c.valor): c
        for c in Conta.objects.filter(usuario=usuario)
    }

    for registro in dados:
        cat_nome = registro["categoria_nome"].upper()

        # Verifica categoria
        if cat_nome in categorias:
            registro["categoria_id"] = categorias[cat_nome].id
            registro["categoria_existe"] = True
        elif cat_nome == "SEM CATEGORIA" or not cat_nome:
            registro["categoria_id"] = None
            registro["categoria_existe"] = True
        else:
            registro["categoria_id"] = None
            registro["categoria_existe"] = False

        # Verifica duplicado
        chave = (registro["data"], registro["descricao"].upper(), registro["valor"])
        if chave in contas_existentes:
            registro["existe_duplicado"] = True
            registro["conta_existente_id"] = contas_existentes[chave].id
        else:
            registro["existe_duplicado"] = False
            registro["conta_existente_id"] = None

    return dados


@transaction.atomic
def importar_dados_relatorio(dados: list, usuario, modo: str = "criar") -> dict:
    """
    Importa os dados do relatório no banco de dados.

    Args:
        dados: Lista de registros a importar (já validados)
        usuario: Usuário dono dos registros
        modo: 'criar' para criar novos, 'sobrescrever' para atualizar existentes

    Returns:
        dict com keys:
            - 'criados': quantidade de registros criados
            - 'atualizados': quantidade de registros atualizados
            - 'ignorados': quantidade de registros ignorados
            - 'erro': mensagem de erro se houver
    """
    criados = 0
    atualizados = 0
    ignorados = 0

    try:
        for registro in dados:
            # Se existe duplicado
            if registro.get("existe_duplicado"):
                if modo == "sobrescrever":
                    # Atualiza registro existente
                    conta = Conta.objects.get(id=registro["conta_existente_id"])
                    conta.tipo = registro["tipo"]
                    conta.descricao = registro["descricao"]
                    conta.valor = registro["valor"]
                    conta.categoria_id = registro.get("categoria_id")
                    conta.transacao_realizada = registro["realizada"]
                    if registro["realizada"] and not conta.data_realizacao:
                        conta.data_realizacao = registro["data"]
                    conta.save()
                    atualizados += 1
                else:
                    # Ignora duplicado
                    ignorados += 1
                continue

            # Cria novo registro
            Conta.objects.create(
                usuario=usuario,
                tipo=registro["tipo"],
                descricao=registro["descricao"],
                valor=registro["valor"],
                data_prevista=registro["data"],
                categoria_id=registro.get("categoria_id"),
                transacao_realizada=registro["realizada"],
                data_realizacao=registro["data"] if registro["realizada"] else None,
            )
            criados += 1

        return {
            "criados": criados,
            "atualizados": atualizados,
            "ignorados": ignorados,
            "erro": None,
        }

    except Exception as e:
        return {
            "criados": criados,
            "atualizados": atualizados,
            "ignorados": ignorados,
            "erro": str(e),
        }


# =====================================================
# FUNÇÕES PARA IMPORTAÇÃO DE INVESTIMENTOS
# =====================================================


def ler_excel_investimentos(arquivo) -> dict:
    """
    Lê a aba 'Investimentos' do arquivo Excel.

    Returns:
        dict com keys:
            - 'dados': lista de dicionários com os ativos
            - 'resumo': dict com total_carteira, quantidade
            - 'erro': mensagem de erro se houver
            - 'tem_aba': bool indicando se a aba existe
    """
    try:
        if hasattr(arquivo, "read"):
            content = arquivo.read()
            arquivo.seek(0)
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        else:
            wb = load_workbook(filename=io.BytesIO(arquivo), read_only=True)

        # Verifica se existe a aba
        if "Investimentos" not in wb.sheetnames:
            wb.close()
            return {"tem_aba": False, "dados": [], "resumo": {}, "erro": None}

        ws = wb["Investimentos"]

        # Encontra cabeçalho (procura por "Ticker")
        header_row = None
        for row_num, row in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), 1
        ):
            if row and row[0] and str(row[0]).strip().upper() == "TICKER":
                header_row = row_num
                break

        if header_row is None:
            wb.close()
            return {
                "tem_aba": True,
                "dados": [],
                "resumo": {},
                "erro": "Aba 'Investimentos' sem cabeçalho válido.",
            }

        # Lê os dados
        dados = []
        total_carteira = Decimal("0.00")

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:
                continue

            # Ignora linhas de resumo
            primeira_col = str(row[0]).strip() if row[0] else ""
            if "TOTAL" in primeira_col.upper():
                continue

            ticker = str(row[0]).strip()
            nome = str(row[1]).strip() if row[1] else ""
            classe = str(row[2]).strip() if row[2] else ""
            categoria = str(row[3]).strip() if row[3] else ""
            subcategoria = str(row[4]).strip() if row[4] else ""
            quantidade = parse_valor(row[5]) or Decimal("0")
            preco_medio = parse_valor(row[6]) or Decimal("0")
            valor_posicao = parse_valor(row[7]) or (quantidade * preco_medio)

            total_carteira += valor_posicao

            dados.append(
                {
                    "ticker": ticker,
                    "nome": nome,
                    "classe": classe,
                    "categoria": categoria,
                    "subcategoria": subcategoria,
                    "quantidade": quantidade,
                    "preco_medio": preco_medio,
                    "valor_posicao": valor_posicao,
                }
            )

        wb.close()

        return {
            "tem_aba": True,
            "dados": dados,
            "resumo": {
                "total_carteira": total_carteira,
                "quantidade": len(dados),
            },
            "erro": None,
        }

    except Exception as e:
        return {"tem_aba": True, "dados": [], "resumo": {}, "erro": str(e)}


def ler_excel_transacoes_investimento(arquivo) -> dict:
    """
    Lê a aba 'Transações Invest.' do arquivo Excel.

    Returns:
        dict com keys:
            - 'dados': lista de dicionários com as transações
            - 'resumo': dict com totais
            - 'erro': mensagem de erro se houver
            - 'tem_aba': bool
    """
    try:
        if hasattr(arquivo, "read"):
            content = arquivo.read()
            arquivo.seek(0)
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        else:
            wb = load_workbook(filename=io.BytesIO(arquivo), read_only=True)

        # Verifica se existe a aba
        if "Transações Invest." not in wb.sheetnames:
            wb.close()
            return {"tem_aba": False, "dados": [], "resumo": {}, "erro": None}

        ws = wb["Transações Invest."]

        # Encontra cabeçalho
        header_row = None
        for row_num, row in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), 1
        ):
            if row and row[0] and str(row[0]).strip().upper() == "DATA":
                header_row = row_num
                break

        if header_row is None:
            wb.close()
            return {
                "tem_aba": True,
                "dados": [],
                "resumo": {},
                "erro": "Aba 'Transações Invest.' sem cabeçalho válido.",
            }

        # Lê os dados
        dados = []
        total_compras = Decimal("0.00")
        total_vendas = Decimal("0.00")
        total_proventos = Decimal("0.00")

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:
                continue

            # Ignora linhas de resumo
            primeira_col = str(row[0]).strip() if row[0] else ""
            if "TOTAL" in primeira_col.upper():
                continue

            # Parse data
            data = row[0]
            if isinstance(data, datetime):
                data = data.date()
            elif isinstance(data, str):
                data = parse_date_br(data)

            if not data:
                continue

            ticker = str(row[1]).strip() if row[1] else ""
            tipo_str = str(row[2]).strip() if row[2] else ""
            quantidade = parse_valor(row[3]) or Decimal("0")
            preco_unitario = parse_valor(row[4]) or Decimal("0")
            taxas = parse_valor(row[5]) or Decimal("0")
            valor_total = parse_valor(row[6]) or Decimal("0")

            # Determina tipo
            tipo_upper = tipo_str.upper()
            if "COMPRA" in tipo_upper:
                tipo = TransacaoInvestimento.TIPO_COMPRA
                total_compras += valor_total
            elif "VENDA" in tipo_upper:
                tipo = TransacaoInvestimento.TIPO_VENDA
                total_vendas += valor_total
            else:
                tipo = TransacaoInvestimento.TIPO_DIVIDENDO
                total_proventos += valor_total

            dados.append(
                {
                    "data": data,
                    "ticker": ticker,
                    "tipo": tipo,
                    "tipo_label": tipo_str,
                    "quantidade": quantidade,
                    "preco_unitario": preco_unitario,
                    "taxas": taxas,
                    "valor_total": valor_total,
                }
            )

        wb.close()

        return {
            "tem_aba": True,
            "dados": dados,
            "resumo": {
                "total_compras": total_compras,
                "total_vendas": total_vendas,
                "total_proventos": total_proventos,
                "quantidade": len(dados),
            },
            "erro": None,
        }

    except Exception as e:
        return {"tem_aba": True, "dados": [], "resumo": {}, "erro": str(e)}


def validar_dados_investimentos(
    dados_ativos: list, dados_transacoes: list, usuario
) -> tuple:
    """
    Valida dados de investimentos contra o banco.

    Adiciona aos registros:
        - 'ativo_existe': bool
        - 'ativo_id': ID do ativo ou None
    """
    # Busca ativos existentes do usuário
    ativos_existentes = {
        a.ticker.upper(): a for a in Ativo.objects.filter(usuario=usuario)
    }

    # Valida ativos
    for ativo in dados_ativos:
        ticker_upper = ativo["ticker"].upper()
        if ticker_upper in ativos_existentes:
            ativo["ativo_existe"] = True
            ativo["ativo_id"] = ativos_existentes[ticker_upper].id
        else:
            ativo["ativo_existe"] = False
            ativo["ativo_id"] = None

    # Valida transações
    for trans in dados_transacoes:
        ticker_upper = trans["ticker"].upper()
        if ticker_upper in ativos_existentes:
            trans["ativo_existe"] = True
            trans["ativo_id"] = ativos_existentes[ticker_upper].id
        else:
            trans["ativo_existe"] = False
            trans["ativo_id"] = None

    return dados_ativos, dados_transacoes


@transaction.atomic
def importar_dados_investimentos(
    dados_ativos: list, dados_transacoes: list, usuario, modo: str = "criar"
) -> dict:
    """
    Importa ativos e transações de investimento.

    Args:
        dados_ativos: Lista de ativos a importar
        dados_transacoes: Lista de transações a importar
        usuario: Usuário dono dos registros
        modo: 'criar' ou 'sobrescrever'

    Returns:
        dict com resultados
    """
    ativos_criados = 0
    ativos_atualizados = 0
    transacoes_criadas = 0
    transacoes_ignoradas = 0

    try:
        # Primeiro, importa/atualiza ativos
        ativos_map = {}  # ticker -> ativo

        for ativo_data in dados_ativos:
            ticker = ativo_data["ticker"].upper()

            if ativo_data.get("ativo_existe"):
                if modo == "sobrescrever":
                    ativo = Ativo.objects.get(id=ativo_data["ativo_id"])
                    ativo.nome = ativo_data["nome"] or ativo.nome
                    ativo.quantidade = ativo_data["quantidade"]
                    ativo.preco_medio = ativo_data["preco_medio"]
                    ativo.save()
                    ativos_atualizados += 1
                    ativos_map[ticker] = ativo
                else:
                    # Mantém o existente
                    ativos_map[ticker] = Ativo.objects.get(id=ativo_data["ativo_id"])
            else:
                # Cria novo ativo
                ativo = Ativo.objects.create(
                    usuario=usuario,
                    ticker=ativo_data["ticker"],
                    nome=ativo_data["nome"],
                    quantidade=ativo_data["quantidade"],
                    preco_medio=ativo_data["preco_medio"],
                )
                ativos_criados += 1
                ativos_map[ticker] = ativo

        # Agora importa transações
        for trans_data in dados_transacoes:
            ticker = trans_data["ticker"].upper()

            # Precisa ter o ativo
            if ticker not in ativos_map:
                transacoes_ignoradas += 1
                continue

            ativo = ativos_map[ticker]

            # Verifica duplicado (mesma data, ticker, tipo, quantidade e valor)
            existe = TransacaoInvestimento.objects.filter(
                usuario=usuario,
                ativo=ativo,
                data=trans_data["data"],
                tipo=trans_data["tipo"],
                quantidade=trans_data["quantidade"],
                valor_total=trans_data["valor_total"],
            ).exists()

            if existe:
                transacoes_ignoradas += 1
                continue

            TransacaoInvestimento.objects.create(
                usuario=usuario,
                ativo=ativo,
                tipo=trans_data["tipo"],
                data=trans_data["data"],
                quantidade=trans_data["quantidade"],
                preco_unitario=trans_data["preco_unitario"],
                taxas=trans_data["taxas"],
                valor_total=trans_data["valor_total"],
            )
            transacoes_criadas += 1

        return {
            "ativos_criados": ativos_criados,
            "ativos_atualizados": ativos_atualizados,
            "transacoes_criadas": transacoes_criadas,
            "transacoes_ignoradas": transacoes_ignoradas,
            "erro": None,
        }

    except Exception as e:
        return {
            "ativos_criados": ativos_criados,
            "ativos_atualizados": ativos_atualizados,
            "transacoes_criadas": transacoes_criadas,
            "transacoes_ignoradas": transacoes_ignoradas,
            "erro": str(e),
        }
