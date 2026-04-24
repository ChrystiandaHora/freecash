"""
Serviço de exportação de relatórios em Excel e PDF.
"""

import io
from datetime import date
from decimal import Decimal

from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend

from django.db.models import Q, Sum, F
from core.models import Conta
from investimento.models import Ativo, Transacao as TransacaoInvestimento


def get_movimentacoes(usuario, data_inicio: date, data_fim: date):
    """
    Busca todas as movimentações do usuário no período.
    Exclui despesas individuais de cartão (apenas faturas consolidadas).
    """
    qs = (
        Conta.objects.filter(
            usuario=usuario,
            data_prevista__gte=data_inicio,
            data_prevista__lte=data_fim,
        )
        .filter(
            # Apenas contas sem cartão OU faturas de cartão
            Q(cartao__isnull=True) | Q(eh_fatura_cartao=True)
        )
        .select_related("categoria", "forma_pagamento", "cartao")
        .order_by("data_prevista", "id")
    )
    return qs


def get_investimentos(usuario, data_inicio: date, data_fim: date):
    """
    Busca ativos do usuário que têm transações no período ou posição > 0.
    """
    # Ativos com posição > 0 ou com transações no período
    ativos_com_transacoes = TransacaoInvestimento.objects.filter(
        usuario=usuario,
        data__gte=data_inicio,
        data__lte=data_fim,
    ).values_list("ativo_id", flat=True)

    qs = (
        Ativo.objects.filter(
            Q(usuario=usuario) & (Q(quantidade__gt=0) | Q(id__in=ativos_com_transacoes))
        )
        .select_related("subcategoria__categoria__classe")
        .order_by("ticker")
    )
    return qs


def get_transacoes_investimento(usuario, data_inicio: date, data_fim: date):
    """
    Busca transações de investimento do usuário no período.
    """
    qs = (
        TransacaoInvestimento.objects.filter(
            usuario=usuario,
            data__gte=data_inicio,
            data__lte=data_fim,
        )
        .select_related("ativo")
        .order_by("data", "id")
    )
    return qs


def get_proventos_data(usuario, data_inicio: date, data_fim: date):
    """
    Busca total de proventos recebidos por ativo no período.
    """
    return (
        TransacaoInvestimento.objects.filter(
            usuario=usuario,
            data__gte=data_inicio,
            data__lte=data_fim,
            tipo=TransacaoInvestimento.TIPO_DIVIDENDO,
        )
        .values("ativo__ticker")
        .annotate(total=Sum("valor_total"))
        .order_by("-total")
    )


def get_alocacao_data(usuario, data_fim: date):
    """
    Busca dados de alocação por classe de ativos.
    """
    ativos = Ativo.objects.filter(usuario=usuario, quantidade__gt=0).select_related(
        "subcategoria__categoria__classe"
    )

    total_portfolio = Decimal("0.00")
    alocacao = {}

    for ativo in ativos:
        valor = ativo.valor_total_atual
        total_portfolio += valor
        classe_nome = (
            ativo.subcategoria.categoria.classe.nome
            if ativo.subcategoria
            else "Outros"
        )
        alocacao[classe_nome] = alocacao.get(classe_nome, Decimal("0.00")) + valor

    # Converter para percentual
    dados = []
    if total_portfolio > 0:
        for classe, valor in alocacao.items():
            percentual = (valor / total_portfolio) * 100
            dados.append({"classe": classe, "valor": valor, "percentual": percentual})

    return sorted(dados, key=lambda x: x["valor"], reverse=True)


def get_comparativo_mensal_data(usuario, data_inicio: date, data_fim: date):
    """
    Busca o saldo (receitas vs despesas) agrupado por mês.
    """
    movs = Conta.objects.filter(
        usuario=usuario,
        data_prevista__gte=data_inicio,
        data_prevista__lte=data_fim,
    ).filter(Q(cartao__isnull=True) | Q(eh_fatura_cartao=True))

    comparativo = {}

    for m in movs:
        mes_ano = m.data_prevista.strftime("%Y-%m")
        if mes_ano not in comparativo:
            comparativo[mes_ano] = {"receitas": Decimal("0"), "despesas": Decimal("0")}

        if m.tipo == Conta.TIPO_RECEITA:
            comparativo[mes_ano]["receitas"] += m.valor
        else:
            comparativo[mes_ano]["despesas"] += m.valor

    # Transformar em lista ordenada
    resultado = []
    for chave in sorted(comparativo.keys()):
        item = comparativo[chave]
        resultado.append(
            {
                "periodo": chave,
                "receitas": item["receitas"],
                "despesas": item["despesas"],
                "saldo": item["receitas"] - item["despesas"],
            }
        )
    return resultado



def gerar_excel(usuario, data_inicio: date, data_fim: date, escopo: str = "completo") -> bytes:
    """
    Gera arquivo Excel com as movimentações, investimentos e transações do período.
    Retorna bytes do arquivo.
    """
    wb = Workbook()
    wb.properties.title = f"Relatório Financeiro {data_inicio.strftime('%d-%m-%Y')} a {data_fim.strftime('%d-%m-%Y')}"
    wb.properties.creator = "FreeCash"

    # Estilos comuns
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_fill_blue = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Remover aba padrão para criar as específicas
    wb.remove(wb.active)

    # =====================
    # ABA: MOVIMENTAÇÕES
    # =====================
    if escopo in ["geral", "completo"]:
        movimentacoes = get_movimentacoes(usuario, data_inicio, data_fim)
        ws = wb.create_sheet("Movimentações")

        # Título
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Relatório de Movimentações - {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Data", "Tipo", "Descrição", "Categoria", "Valor (R$)", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border

        total_receitas = total_despesas = Decimal("0.00")
        for row_num, mov in enumerate(movimentacoes, 4):
            tipo_label = "Receita" if mov.tipo == Conta.TIPO_RECEITA else "Despesa"
            if mov.tipo == Conta.TIPO_RECEITA: total_receitas += mov.valor
            else: total_despesas += mov.valor

            data = [mov.data_prevista.strftime("%d/%m/%Y"), tipo_label, mov.descricao, 
                    mov.categoria.nome if mov.categoria else "Sem cat.", float(mov.valor), 
                    "Realizada" if mov.transacao_realizada else "Pendente"]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if col == 5: cell.number_format = "#,##0.00"; cell.alignment = Alignment(horizontal="right")

        # Resumo Mensal (Comparativo) na mesma aba ou nova? Vamos fazer nova aba por organização
        comp_data = get_comparativo_mensal_data(usuario, data_inicio, data_fim)
        ws_comp = wb.create_sheet("Resumo Mensal")
        ws_comp.append(["Mês/Ano", "Receitas", "Despesas", "Saldo"])
        for c in ws_comp[1]: c.font = header_font; c.fill = header_fill; c.border = thin_border
        for item in comp_data:
            ws_comp.append([item["periodo"], float(item["receitas"]), float(item["despesas"]), float(item["saldo"])])
        for row in ws_comp.iter_rows(min_row=2):
            for cell in row[1:]: cell.number_format = "#,##0.00"

    # =====================
    # SEÇÃO: INVESTIMENTOS
    # =====================
    if escopo in ["investimentos", "completo"]:
        investimentos = get_investimentos(usuario, data_inicio, data_fim)
        ws_inv = wb.create_sheet("Carteira")
        invest_headers = ["Ticker", "Nome", "Classe", "Categoria", "Quantidade", "P. Médio", "Investido", "Mercado", "Meta (%)", "Valor Ideal", "Sugestão", "Lucro/Prej."]
        for col, h in enumerate(invest_headers, 1):
            cell = ws_inv.cell(row=1, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill_blue; cell.border = thin_border

        total_portfolio = sum(a.valor_total_atual for a in investimentos)

        for row_num, ativo in enumerate(investimentos, 2):
            val_inv = ativo.valor_investido
            val_mer = ativo.valor_total_atual
            meta = ativo.meta_porcentagem
            val_ideal = (meta / 100) * total_portfolio if total_portfolio > 0 else 0
            sugestao = val_ideal - val_mer
            
            data = [
                ativo.ticker, ativo.nome or "", 
                ativo.subcategoria.categoria.classe.nome if ativo.subcategoria else "",
                ativo.subcategoria.categoria.nome if ativo.subcategoria else "",
                float(ativo.quantidade), float(ativo.preco_medio), float(val_inv), 
                float(val_mer), float(meta), float(val_ideal), float(sugestao), float(val_mer - val_inv)
            ]
            for col, val in enumerate(data, 1):
                cell = ws_inv.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if col >= 5: cell.number_format = "#,##0.00"
                if col == 9: cell.number_format = "0.00\"%\""

        # Aba de Proventos
        proventos = get_proventos_data(usuario, data_inicio, data_fim)
        ws_prov = wb.create_sheet("Proventos")
        ws_prov.append(["Ticker", "Total Recebido (R$)"])
        for c in ws_prov[1]: c.font = header_font; c.fill = header_fill_blue; c.border = thin_border
        for p in proventos:
            ws_prov.append([p["ativo__ticker"], float(p["total"])])
            ws_prov.cell(row=ws_prov.max_row, column=2).number_format = "#,##0.00"

        # Aba de Alocação
        aloc = get_alocacao_data(usuario, data_fim)
        ws_aloc = wb.create_sheet("Alocação")
        ws_aloc.append(["Classe", "Valor (R$)", "Percentual (%)"])
        for c in ws_aloc[1]: c.font = header_font; c.fill = header_fill_blue; c.border = thin_border
        for a in aloc:
            ws_aloc.append([a["classe"], float(a["valor"]), float(a["percentual"])])
            ws_aloc.cell(row=ws_aloc.max_row, column=2).number_format = "#,##0.00"
            ws_aloc.cell(row=ws_aloc.max_row, column=3).number_format = "0.00\"%\""

        # Aba de Transações
        transacoes_invest = get_transacoes_investimento(usuario, data_inicio, data_fim)
        ws_tr = wb.create_sheet("Transações Invest.")
        ws_tr.append(["Data", "Ticker", "Tipo", "Qtd", "Preço", "Taxas", "Total"])
        for c in ws_tr[1]: c.font = header_font; c.fill = header_fill_blue; c.border = thin_border
        for t in transacoes_invest:
            ws_tr.append([t.data.strftime("%d/%m/%Y"), t.ativo.ticker, t.get_tipo_display(), float(t.quantidade), float(t.preco_unitario), float(t.taxas), float(t.valor_total)])
            for col in range(4, 8): ws_tr.cell(row=ws_tr.max_row, column=col).number_format = "#,##0.00"

    # Salvar em bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_grafico_alocacao(alocacao_dados):
    """
    Gera um objeto Drawing com gráfico de pizza para alocação.
    """
    if not alocacao_dados:
        return None

    d = Drawing(400, 200)
    pc = Pie()
    pc.x = 150
    pc.y = 50
    pc.width = 120
    pc.height = 120
    
    pc.data = [float(x["valor"]) for x in alocacao_dados]
    pc.labels = [x["classe"] for x in alocacao_dados]
    
    # Cores bonitas
    pc.slices.strokeWidth = 0.5
    cores = [colors.HexColor("#10B981"), colors.HexColor("#3B82F6"), colors.HexColor("#8B5CF6"), 
             colors.HexColor("#F59E0B"), colors.HexColor("#EF4444"), colors.HexColor("#6B7280")]
    
    for i, color in enumerate(cores):
        if i < len(pc.data):
            pc.slices[i].fillColor = color

    d.add(pc)
    
    # Legenda
    legend = Legend()
    legend.x = 300
    legend.y = 130
    legend.dx = 8
    legend.dy = 8
    legend.fontName = 'Helvetica'
    legend.fontSize = 7
    legend.columnMaximum = 10
    legend.alignment = 'right'
    legend.colorNamePairs = [(cores[i % len(cores)], f"{x['classe']} ({x['percentual']:.1f}%)") for i, x in enumerate(alocacao_dados)]
    d.add(legend)
    
    return d


def gerar_pdf(usuario, data_inicio: date, data_fim: date, escopo: str = "completo") -> bytes:
    """
    Gera arquivo PDF com as movimentações, investimentos e transações do período.
    Retorna bytes do arquivo.
    """
    movimentacoes = get_movimentacoes(usuario, data_inicio, data_fim)
    investimentos = get_investimentos(usuario, data_inicio, data_fim)
    transacoes_invest = get_transacoes_investimento(usuario, data_inicio, data_fim)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Relatório Financeiro {data_inicio.strftime('%d-%m-%Y')} a {data_fim.strftime('%d-%m-%Y')}",
        author="FreeCash",
    )

    elements = []
    styles = getSampleStyleSheet()

    # Estilo de título
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=20,
        alignment=1,  # Center
    )

    # Estilo de seção
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=10,
        spaceBefore=15,
        textColor=colors.HexColor("#1F2937"),
    )

    # =====================
    # SEÇÃO 1: MOVIMENTAÇÕES
    # =====================
    title = Paragraph(
        f"Relatório Financeiro<br/>{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
        title_style,
    )
    elements.append(title)
    elements.append(Spacer(1, 5 * mm))

    if escopo in ["geral", "completo"]:
        section_mov = Paragraph("📊 Movimentações", section_style)
        elements.append(section_mov)

        # Tabela de movimentações
        table_data = [["Data", "Tipo", "Descrição", "Categoria", "Valor", "Status"]]

        total_receitas = Decimal("0.00")
        total_despesas = Decimal("0.00")

        for mov in movimentacoes:
            tipo_label = "Receita" if mov.tipo == Conta.TIPO_RECEITA else "Despesa"
            categoria_nome = mov.categoria.nome if mov.categoria else "Sem cat."
            status = "OK" if mov.transacao_realizada else "Pend."

            if mov.tipo == Conta.TIPO_RECEITA: total_receitas += mov.valor
            else: total_despesas += mov.valor

            desc = mov.descricao[:30] + "..." if len(mov.descricao) > 30 else mov.descricao
            table_data.append([mov.data_prevista.strftime("%d/%m/%Y"), tipo_label, desc, categoria_nome[:15], 
                               f"R$ {mov.valor:,.2f}", status])

        col_widths = [50, 60, 180, 100, 85, 60]
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#10B981")),
            ("ALIGN", (4, 1), (4, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 5 * mm))

        # Novo: Comparativo Mensal
        elements.append(Paragraph("📅 Comparativo Mensal", section_style))
        comp_data = get_comparativo_mensal_data(usuario, data_inicio, data_fim)
        comp_table_data = [["Período", "Receitas", "Despesas", "Saldo"]]
        
        t_rec = t_des = Decimal("0.00")
        for c in comp_data:
            t_rec += c["receitas"]
            t_des += c["despesas"]
            comp_table_data.append([c["periodo"], f"R$ {c['receitas']:,.2f}", f"R$ {c['despesas']:,.2f}", f"R$ {c['saldo']:,.2f}"])
        
        # Linha de Total Geral
        comp_table_data.append(["TOTAL GERAL", f"R$ {t_rec:,.2f}", f"R$ {t_des:,.2f}", f"R$ {t_rec - t_des:,.2f}"])
        
        comp_table = Table(comp_table_data, colWidths=[100, 100, 100, 100])
        comp_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.grey),
            ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("TEXTCOLOR", (3, -1), (3, -1), colors.HexColor("#10B981") if (t_rec - t_des) >= 0 else colors.red),
        ]))
        elements.append(comp_table)
        elements.append(Spacer(1, 10 * mm))

    # =====================
    # SEÇÃO 2: INVESTIMENTOS
    # =====================
    if escopo in ["investimentos", "completo"] and investimentos.exists():
        elements.append(PageBreak())
        section_invest = Paragraph("📈 Carteira de Investimentos", section_style)
        elements.append(section_invest)

        # Gráfico de Alocação
        elements.append(Paragraph("Alocação por Classe de Ativo", styles["Normal"]))
        aloc_dados = get_alocacao_data(usuario, data_fim)
        grafico = render_grafico_alocacao(aloc_dados)
        if grafico:
            elements.append(grafico)
            elements.append(Spacer(1, 5 * mm))

        invest_data = [["Ticker", "Classe", "Qtd", "PM", "Mercado", "Meta (%)", "Ideal", "Lucro/P"]]
        total_mer = sum(a.valor_total_atual for a in investimentos)

        for ativo in investimentos:
            vm = ativo.valor_total_atual
            meta = ativo.meta_porcentagem
            val_ideal = (meta / 100) * total_mer if total_mer > 0 else 0
            classe = ativo.subcategoria.categoria.classe.nome if ativo.subcategoria else ""
            invest_data.append([
                ativo.ticker, 
                classe[:10], 
                f"{ativo.quantidade:,.2f}", 
                f"R$ {ativo.preco_medio:,.2f}", 
                f"R$ {vm:,.2f}", 
                f"{meta:,.2f}%",
                f"R$ {val_ideal:,.2f}",
                f"R$ {vm - ativo.valor_investido:,.2f}"
            ])

        # Ajuste de larguras: total ~520 pontos para caber no A4
        table_inv = Table(invest_data, colWidths=[55, 65, 55, 75, 80, 50, 80, 80])
        table_inv.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#3B82F6")),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),  # Fonte levemente menor para caber tudo
        ]))
        elements.append(table_inv)
        elements.append(Spacer(1, 10 * mm))

        # Novo: Resumo de Proventos
        elements.append(Paragraph("💰 Proventos Recebidos no Período", section_style))
        prov_dados = get_proventos_data(usuario, data_inicio, data_fim)
        if prov_dados.exists():
            prov_table_data = [["Ticker", "Total Recebido"]]
            for p in prov_dados:
                prov_table_data.append([p["ativo__ticker"], f"R$ {p['total']:,.2f}"])
            
            table_prov = Table(prov_table_data, colWidths=[100, 150])
            table_prov.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#8B5CF6")),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]))
            elements.append(table_prov)
        else:
            elements.append(Paragraph("Nenhum provento recebido no período.", styles["Normal"]))
        
        elements.append(Spacer(1, 10 * mm))

        # Transações de Investimento
        if transacoes_invest.exists():
            elements.append(Paragraph("Transactions Tracking", section_style))
            tr_data = [["Data", "Ticker", "Tipo", "Qtd", "Preço", "Total"]]
            for t in transacoes_invest:
                tr_data.append([t.data.strftime("%d/%m/%Y"), t.ativo.ticker, t.get_tipo_display()[:10], 
                                f"{t.quantidade:,.2f}", f"R$ {t.preco_unitario:,.2f}", f"R$ {t.valor_total:,.2f}"])
            
            table_tr = Table(tr_data, colWidths=[60, 70, 90, 80, 100, 100])
            table_tr.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#8B5CF6")),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            elements.append(table_tr)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
