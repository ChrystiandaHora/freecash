from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
import io
from core.models import Categoria, Conta, FormaPagamento, ConfigUsuario
from core.services.encryption import encrypt_to_zip


def gerar_backup_excel(usuario, password=None):
    wb = Workbook()

    # =========================
    # Aba 0: Metadata
    # =========================
    ws_meta = wb.active
    ws_meta.title = "metadata"
    ws_meta.append(["chave", "valor"])
    ws_meta.append(["backup_version", "2"])  # Version 2 uses UUIDs
    ws_meta.append(["generated_at", timezone.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_meta.append(
        ["usuario_uuid", str(usuario.config.uuid) if hasattr(usuario, "config") else ""]
    )
    ws_meta.append(["username", usuario.get_username()])

    # =========================
    # Aba 1: Contas
    # =========================
    ws_contas = wb.create_sheet("contas")
    ws_contas.append(
        [
            "uuid",
            "tipo",
            "descricao",
            "valor",
            "data_prevista",
            "transacao_realizada",
            "data_realizacao",
            "categoria_uuid",
            "forma_pagamento_uuid",
            "is_legacy",
            "origem_ano",
            "origem_mes",
            "origem_linha",
            "criada_em",
            "atualizada_em",
        ]
    )

    contas = (
        Conta.objects.filter(usuario=usuario)
        .select_related("categoria", "forma_pagamento")
        .order_by("data_prevista", "id")
    )

    for c in contas:
        ws_contas.append(
            [
                str(c.uuid),
                c.tipo,
                c.descricao or "",
                float(c.valor),
                c.data_prevista.strftime("%Y-%m-%d") if c.data_prevista else "",
                bool(c.transacao_realizada),
                c.data_realizacao.strftime("%Y-%m-%d") if c.data_realizacao else "",
                str(c.categoria.uuid) if c.categoria else "",
                str(c.forma_pagamento.uuid) if c.forma_pagamento else "",
                bool(getattr(c, "is_legacy", False)),
                getattr(c, "origem_ano", None),
                getattr(c, "origem_mes", None),
                getattr(c, "origem_linha", "") or "",
                c.criada_em.strftime("%Y-%m-%d %H:%M:%S") if c.criada_em else "",
                c.atualizada_em.strftime("%Y-%m-%d %H:%M:%S")
                if c.atualizada_em
                else "",
            ]
        )

    # =========================
    # Aba 2: Categorias
    # =========================
    ws_cat = wb.create_sheet("categorias")
    ws_cat.append(["uuid", "nome", "tipo", "is_default", "criada_em", "atualizada_em"])

    for cat in Categoria.objects.filter(usuario=usuario).order_by("id"):
        ws_cat.append(
            [
                str(cat.uuid),
                cat.nome,
                cat.tipo,
                bool(getattr(cat, "is_default", False)),
                cat.criada_em.strftime("%Y-%m-%d %H:%M:%S") if cat.criada_em else "",
                cat.atualizada_em.strftime("%Y-%m-%d %H:%M:%S")
                if cat.atualizada_em
                else "",
            ]
        )

    # =========================
    # Aba 3: Formas de pagamento
    # =========================
    ws_fp = wb.create_sheet("formas_pagamento")
    ws_fp.append(["uuid", "nome", "ativa", "criada_em", "atualizada_em"])

    for fp in FormaPagamento.objects.filter(usuario=usuario).order_by("id"):
        ws_fp.append(
            [
                str(fp.uuid),
                fp.nome,
                bool(getattr(fp, "ativa", True)),
                fp.criada_em.strftime("%Y-%m-%d %H:%M:%S") if fp.criada_em else "",
                fp.atualizada_em.strftime("%Y-%m-%d %H:%M:%S")
                if fp.atualizada_em
                else "",
            ]
        )

    # =========================
    # Aba 5: Configurações
    # =========================
    ws_conf = wb.create_sheet("configuracoes")
    ws_conf.append(["uuid", "moeda_padrao", "ultimo_export_em"])

    config = ConfigUsuario.objects.filter(usuario=usuario).first()
    if config:
        ws_conf.append(
            [
                str(config.uuid),
                getattr(config, "moeda_padrao", "BRL"),
                config.ultimo_export_em.strftime("%Y-%m-%d %H:%M:%S")
                if config.ultimo_export_em
                else "",
            ]
        )

    # =========================
    # Finalize
    # =========================

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    file_data = excel_buffer.getvalue()

    filename_base = f"backup_freecash_{usuario.get_username()}_{timezone.now().strftime('%Y%m%d_%H%M%S')}"

    if password:
        # Encrypt
        encrypted_data = encrypt_to_zip(f"{filename_base}.xlsx", file_data, password)
        response = HttpResponse(encrypted_data, content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.zip"'
    else:
        # Standard XLSX
        response = HttpResponse(
            file_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'

    # Update config
    if config:
        config.ultimo_export_em = timezone.now()
        config.save(update_fields=["ultimo_export_em"])

    return response
