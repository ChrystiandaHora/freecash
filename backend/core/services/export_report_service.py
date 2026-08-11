"""
Serviço de exportação de relatórios em Excel e PDF.
"""

import csv
import io
import math
from datetime import date
from decimal import Decimal
from xml.sax.saxutils import escape as escape_xml

from django.db.models import Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)
from reportlab.graphics.shapes import Drawing, Line, Rect, String
from reportlab.graphics.charts.doughnut import Doughnut
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.legends import Legend

from core.models import Conta
from investimento.models import Ativo, Transacao as TransacaoInvestimento


# =============================================================================
# IDENTIDADE VISUAL
# =============================================================================

#: Paleta institucional do relatório, alinhada ao azul da marca usado no frontend.
PALETA = {
    "marca": "#007ACC",
    "marca_esc": "#005B99",
    "tinta": "#1F2937",
    "tinta_sec": "#6B7280",
    "linha": "#E5E7EB",
    "zebra": "#F7F9FB",
    "destaque": "#EEF6FC",
    "positivo": "#059669",
    "negativo": "#DC2626",
}

#: Sequência de cores para séries categóricas (fatias de gráfico, classes de ativo).
SERIE_CORES = [
    "#007ACC", "#00A3A3", "#7C3AED", "#F59E0B", "#DC2626",
    "#0EA5E9", "#65A30D", "#DB2777", "#6B7280", "#B45309",
]

MESES_ABREV = (
    "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez",
)

MARGEM_LATERAL = 16 * mm
MARGEM_TOPO = 22 * mm
MARGEM_BASE = 18 * mm
RODAPE_INSTITUCIONAL = "FreeCash · Sistema de Gestão Financeira Pessoal · Uso interno"
MESES_NO_GRAFICO = 12


def cor(chave: str):
    """Converte uma chave da paleta institucional em uma cor do reportlab.

    Args:
        chave (str): Nome da cor em PALETA (ex.: 'marca', 'negativo').

    Returns:
        Color: Instância de cor do reportlab correspondente ao hexadecimal.
    """
    return colors.HexColor(PALETA[chave])


def cor_serie(indice: int):
    """Retorna a cor da série categórica na posição informada, ciclando a paleta.

    Args:
        indice (int): Posição da série (0 para a primeira).

    Returns:
        Color: Cor do reportlab, garantindo cor definida para qualquer quantidade de itens.
    """
    return colors.HexColor(SERIE_CORES[indice % len(SERIE_CORES)])


# =============================================================================
# FORMATAÇÃO PT-BR
# =============================================================================


def _dec(valor) -> Decimal:
    """Normaliza qualquer entrada numérica para Decimal, tratando None como zero.

    Args:
        valor: Valor numérico (Decimal, int, float, str) ou None.

    Returns:
        Decimal: Valor convertido, ou Decimal("0") quando a entrada é nula/inválida.
    """
    if valor is None:
        return Decimal("0")
    if isinstance(valor, Decimal):
        return valor
    try:
        return Decimal(str(valor))
    except (ArithmeticError, ValueError):
        return Decimal("0")


def formatar_numero(valor, casas: int = 2) -> str:
    """Formata um número no padrão brasileiro (ponto de milhar, vírgula decimal).

    Não depende de `locale`, que não é confiável em containers enxutos.

    Args:
        valor: Valor numérico a ser formatado.
        casas (int, optional): Quantidade de casas decimais. Defaults to 2.

    Returns:
        str: Número formatado, por exemplo "3.048,21".
    """
    texto = f"{_dec(valor):,.{casas}f}"
    return texto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def formatar_moeda(valor, simbolo: bool = True) -> str:
    """Formata um valor monetário no padrão brasileiro com sinal antes do símbolo.

    Args:
        valor: Valor monetário a ser formatado.
        simbolo (bool, optional): Inclui o prefixo "R$ ". Defaults to True.

    Returns:
        str: Valor formatado, por exemplo "R$ 3.048,21" ou "-R$ 250,00".
    """
    valor_decimal = _dec(valor)
    sinal = "-" if valor_decimal < 0 else ""
    corpo = formatar_numero(abs(valor_decimal))
    return f"{sinal}R$ {corpo}" if simbolo else f"{sinal}{corpo}"


def formatar_percentual(valor, casas: int = 1) -> str:
    """Formata um percentual no padrão brasileiro.

    Args:
        valor: Valor percentual (já em escala 0-100).
        casas (int, optional): Casas decimais. Defaults to 1.

    Returns:
        str: Percentual formatado, por exemplo "70,1%".
    """
    return f"{formatar_numero(valor, casas)}%"


def formatar_mes_ano(chave: str) -> str:
    """Converte a chave de período 'YYYY-MM' na forma legível 'mmm/YYYY'.

    Args:
        chave (str): Período no formato "2026-01".

    Returns:
        str: Período legível, por exemplo "jan/2026". Devolve a entrada quando inválida.
    """
    try:
        ano, mes = str(chave).split("-")[:2]
        return f"{MESES_ABREV[int(mes) - 1]}/{ano}"
    except (ValueError, IndexError, TypeError):
        return str(chave)


def _teto_eixo(valor) -> float:
    """Arredonda um valor para cima até uma referência "redonda" para o topo do eixo.

    Args:
        valor: Maior valor da série plotada.

    Returns:
        float: Limite superior do eixo, sempre positivo.
    """
    maximo = float(_dec(valor))
    if maximo <= 0:
        return 1.0
    expoente = math.floor(math.log10(maximo))
    base = 10.0 ** expoente
    for multiplo in (1, 2, 2.5, 5, 10):
        if maximo <= multiplo * base:
            return multiplo * base
    return 10 * base


# =============================================================================
# CONSULTAS DE DADOS
# =============================================================================


def get_movimentacoes(usuario, data_inicio: date, data_fim: date):
    """Busca todas as movimentações gerais do usuário no período informado.

    Exclui despesas de cartão individuais para evitar duplicidade de lançamentos,
    retornando apenas contas de caixa (sem cartão associado) e faturas de cartão consolidadas.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_inicio (date): Limite inferior do período de busca.
        data_fim (date): Limite superior do período de busca.

    Returns:
        QuerySet: Lista de lançamentos de Conta ordenados por data prevista e id.
    """
    qs = (
        Conta.objects.filter(
            usuario=usuario,
            data_prevista__gte=data_inicio,
            data_prevista__lte=data_fim,
        )
        .filter(
            # Apenas contas sem cartão OU faturas de cartão
            Q(cartao__isnull=True) | Q(eh_fatura_cartao=True)
        )
        .select_related("categoria", "cartao")
        .order_by("data_prevista", "id")
    )
    return qs


def get_investimentos(usuario, data_inicio: date, data_fim: date):
    """Busca os ativos do usuário que possuam posição ativa ou transações no período.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_inicio (date): Limite de início do período.
        data_fim (date): Limite de fim do período.

    Returns:
        QuerySet: Filtro de ativos B3 ordenados pelo ticker alfabeticamente.
    """
    # Ativos com posição > 0 ou com transações no período
    ativos_com_transacoes = TransacaoInvestimento.objects.filter(
        usuario=usuario,
        data__gte=data_inicio,
        data__lte=data_fim,
    ).values_list("ativo_id", flat=True)

    qs = (
        Ativo.objects.filter(
            Q(usuario=usuario) & (Q(quantidade__gt=0) | Q(id__in=ativos_com_transacoes))
        )
        .select_related("subcategoria__categoria__classe")
        .order_by("ticker")
    )
    return qs


def get_transacoes_investimento(usuario, data_inicio: date, data_fim: date):
    """Busca o histórico de ordens de compra/venda de investimentos no período.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_inicio (date): Data de início do período de movimentações.
        data_fim (date): Data final do período de movimentações.

    Returns:
        QuerySet: Histórico de ordens executadas ordenadas por data e id.
    """
    qs = (
        TransacaoInvestimento.objects.filter(
            usuario=usuario,
            data__gte=data_inicio,
            data__lte=data_fim,
        )
        .select_related("ativo")
        .order_by("data", "id")
    )
    return qs


def get_proventos_data(usuario, data_inicio: date, data_fim: date):
    """Agrupa e soma o total de dividendos/juros recebidos por ativo no período.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_inicio (date): Data de início da apuração.
        data_fim (date): Data de fim da apuração.

    Returns:
        QuerySet: Agrupado por ticker contendo o somatório dos proventos recebidos.
    """
    return (
        TransacaoInvestimento.objects.filter(
            usuario=usuario,
            data__gte=data_inicio,
            data__lte=data_fim,
            tipo=TransacaoInvestimento.TIPO_DIVIDENDO,
        )
        .values("ativo__ticker")
        .annotate(total=Sum("valor_total"))
        .order_by("-total")
    )


def get_despesas_por_categoria(usuario, data_inicio: date, data_fim: date, limite: int = 10):
    """Agrupa as despesas do período por categoria, ordenadas da maior para a menor.

    Aplica o mesmo filtro anti-duplicidade de cartão usado em `get_movimentacoes`.
    Quando existem mais categorias que o limite, o excedente é somado em "Outras
    categorias" para que o total continue fechando com o relatório.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_inicio (date): Limite inferior do período.
        data_fim (date): Limite superior do período.
        limite (int, optional): Quantidade máxima de categorias detalhadas. Defaults to 10.

    Returns:
        list[dict]: Dicionários com 'categoria', 'total' e 'percentual' sobre o total de despesas.
    """
    agregado = (
        Conta.objects.filter(
            usuario=usuario,
            tipo=Conta.TIPO_DESPESA,
            data_prevista__gte=data_inicio,
            data_prevista__lte=data_fim,
        )
        .filter(Q(cartao__isnull=True) | Q(eh_fatura_cartao=True))
        .values("categoria__nome")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )

    itens = [
        {
            "categoria": linha["categoria__nome"] or "Sem categoria",
            "total": _dec(linha["total"]),
        }
        for linha in agregado
    ]

    total_despesas = sum((item["total"] for item in itens), Decimal("0.00"))

    if limite and len(itens) > limite:
        excedente = sum((item["total"] for item in itens[limite:]), Decimal("0.00"))
        itens = itens[:limite]
        itens.append({"categoria": "Outras categorias", "total": excedente})

    for item in itens:
        item["percentual"] = (
            (item["total"] / total_despesas * 100) if total_despesas > 0 else Decimal("0")
        )

    return itens


def get_alocacao_data(usuario, data_fim: date):
    """Gera a distribuição percentual e absoluta do portfólio por classe de ativos.

    Analisa a custódia total do investidor na data limite e agrupa os valores sob
    as classes de Renda Fixa, Ações, FIIs, etc.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_fim (date): Data limite para consideração de saldo na custódia.

    Returns:
        list[dict]: Lista de dicionários ordenada com 'classe', 'valor' e 'percentual'.
    """
    ativos = Ativo.objects.filter(usuario=usuario, quantidade__gt=0).select_related(
        "subcategoria__categoria__classe"
    )

    total_portfolio = Decimal("0.00")
    alocacao = {}

    for ativo in ativos:
        valor = ativo.valor_total_atual
        total_portfolio += valor
        classe_nome = (
            ativo.subcategoria.categoria.classe.nome
            if ativo.subcategoria
            else "Outros"
        )
        alocacao[classe_nome] = alocacao.get(classe_nome, Decimal("0.00")) + valor

    # Converter para percentual
    dados = []
    if total_portfolio > 0:
        for classe, valor in alocacao.items():
            percentual = (valor / total_portfolio) * 100
            dados.append({"classe": classe, "valor": valor, "percentual": percentual})

    return sorted(dados, key=lambda x: x["valor"], reverse=True)


def get_comparativo_mensal_data(usuario, data_inicio: date, data_fim: date):
    """Calcula mensalmente o saldo total apurado (receitas menos despesas).

    Agrupa os totais de fluxo de caixa por chaves no formato YYYY-MM para o período
    selecionado.

    Args:
        usuario (User): Instância do usuário Django proprietário.
        data_inicio (date): Limite inferior do período.
        data_fim (date): Limite superior do período.

    Returns:
        list[dict]: Lista contendo dicionários com 'periodo', 'receitas', 'despesas' e 'saldo'.
    """
    movs = Conta.objects.filter(
        usuario=usuario,
        data_prevista__gte=data_inicio,
        data_prevista__lte=data_fim,
    ).filter(Q(cartao__isnull=True) | Q(eh_fatura_cartao=True))

    comparativo = {}

    for m in movs:
        mes_ano = m.data_prevista.strftime("%Y-%m")
        if mes_ano not in comparativo:
            comparativo[mes_ano] = {"receitas": Decimal("0"), "despesas": Decimal("0")}

        if m.tipo == Conta.TIPO_RECEITA:
            comparativo[mes_ano]["receitas"] += m.valor
        else:
            comparativo[mes_ano]["despesas"] += m.valor

    # Transformar em lista ordenada
    resultado = []
    for chave in sorted(comparativo.keys()):
        item = comparativo[chave]
        resultado.append(
            {
                "periodo": chave,
                "receitas": item["receitas"],
                "despesas": item["despesas"],
                "saldo": item["receitas"] - item["despesas"],
            }
        )
    return resultado




# =============================================================================
# PLANILHAS — FORMATOS E COMPONENTES
# =============================================================================

#: Formatos numéricos do Excel. Os separadores são resolvidos pelo próprio Excel
#: conforme o idioma do usuário, então "#,##0.00" já aparece como 1.234,56 em pt-BR.
FORMATO_MOEDA = "R$ #,##0.00;[Red]-R$ #,##0.00"
FORMATO_QUANTIDADE = "#,##0.00####"
FORMATO_PERCENTUAL = "0.0%"
FORMATO_DATA = "dd/mm/yyyy"
FORMATO_MES = "mmm/yyyy"

_LARGURA_MINIMA = 11
_LARGURA_MAXIMA = 44


def _largura_visual(valor, formato) -> int:
    """Estima em caracteres o espaço ocupado por uma célula depois de formatada.

    Args:
        valor: Conteúdo da célula.
        formato (str | None): Formato numérico aplicado à coluna.

    Returns:
        int: Quantidade estimada de caracteres exibidos.
    """
    if valor is None or valor == "":
        return 0
    if isinstance(valor, date):
        return 10 if formato == FORMATO_DATA else 9
    if isinstance(valor, (int, float, Decimal)) and not isinstance(valor, bool):
        if formato == FORMATO_PERCENTUAL:
            return len(formatar_percentual(_dec(valor) * 100))
        if formato == FORMATO_MOEDA:
            return len(formatar_moeda(valor))
        return len(formatar_numero(valor))
    return len(str(valor))


def _montar_aba(wb, nome, cabecalhos, linhas, formatos, cor_aba, linha_total=None):
    """Cria uma aba já formatada, com o mesmo padrão visual em todo o arquivo.

    Aplica cabeçalho colorido congelado, filtro automático sobre os dados,
    formatos numéricos por coluna, largura de coluna calculada pelo conteúdo e
    uma linha de totalização opcional (fora do filtro, para não ser reordenada).

    Args:
        wb (Workbook): Pasta de trabalho de destino.
        nome (str): Nome da aba.
        cabecalhos (Sequence[str]): Títulos das colunas.
        linhas (Sequence[Sequence]): Linhas de dados.
        formatos (Sequence[str | None]): Formato numérico de cada coluna (None = texto).
        cor_aba (str): Cor hexadecimal da guia da aba.
        linha_total (Sequence | None): Linha de totalização, se houver.

    Returns:
        Worksheet: A aba criada.
    """
    ws = wb.create_sheet(nome)
    ws.sheet_properties.tabColor = cor_aba.lstrip("#")

    preenchimento = PatternFill(
        "solid", start_color=PALETA["marca_esc"][1:], end_color=PALETA["marca_esc"][1:]
    )
    for coluna, titulo in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=coluna, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = preenchimento
        celula.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for linha in linhas:
        ws.append(list(linha))

    ultima_linha = ws.max_row
    for coluna, formato in enumerate(formatos, 1):
        if not formato:
            continue
        for numero in range(2, ultima_linha + 1):
            celula = ws.cell(row=numero, column=coluna)
            celula.number_format = formato
            celula.alignment = Alignment(horizontal="right")

    if linhas:
        ultima_coluna = get_column_letter(len(cabecalhos))
        ws.auto_filter.ref = f"A1:{ultima_coluna}{ultima_linha}"

    if linha_total:
        ws.append(list(linha_total))
        borda = Border(top=Side(style="thin", color=PALETA["marca"][1:]))
        destaque = PatternFill(
            "solid", start_color=PALETA["destaque"][1:], end_color=PALETA["destaque"][1:]
        )
        for coluna in range(1, len(cabecalhos) + 1):
            celula = ws.cell(row=ws.max_row, column=coluna)
            celula.font = Font(bold=True)
            celula.border = borda
            celula.fill = destaque
            formato = formatos[coluna - 1] if coluna - 1 < len(formatos) else None
            # O formato só vale para células numéricas: a linha de total costuma
            # trazer um rótulo de texto em colunas de data/moeda.
            if formato and isinstance(celula.value, (int, float, Decimal)):
                celula.number_format = formato
                celula.alignment = Alignment(horizontal="right")

    matriz = [list(cabecalhos)] + [list(linha) for linha in linhas]
    if linha_total:
        matriz.append(list(linha_total))
    for coluna in range(1, len(cabecalhos) + 1):
        formato = formatos[coluna - 1] if coluna - 1 < len(formatos) else None
        largura = max(
            _largura_visual(linha[coluna - 1] if coluna - 1 < len(linha) else None, formato)
            for linha in matriz
        )
        ws.column_dimensions[get_column_letter(coluna)].width = min(
            max(largura + 3, _LARGURA_MINIMA), _LARGURA_MAXIMA
        )
    return ws


def _rotulo_escopo(escopo: str) -> str:
    """Descreve em português o conteúdo coberto pelo escopo informado.

    Args:
        escopo (str): 'geral', 'investimentos' ou 'completo'.

    Returns:
        str: Descrição legível do escopo.
    """
    return {
        "geral": "Movimentações gerais",
        "investimentos": "Carteira de investimentos",
        "completo": "Movimentações gerais e carteira de investimentos",
    }.get(escopo, "Movimentações gerais e carteira de investimentos")


def _aba_resumo(wb, usuario, data_inicio, data_fim, escopo, indicadores) -> None:
    """Cria a primeira aba, com a identificação do relatório e os indicadores.

    Args:
        wb (Workbook): Pasta de trabalho de destino.
        usuario (User): Usuário solicitante.
        data_inicio (date): Início do período.
        data_fim (date): Fim do período.
        escopo (str): Escopo do relatório.
        indicadores (Sequence[tuple]): Pares (rótulo, valor) já na ordem de exibição.
    """
    ws = wb.create_sheet("Resumo", 0)
    ws.sheet_properties.tabColor = PALETA["marca"][1:]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 34

    ws["A1"] = "Relatório Financeiro"
    ws["A1"].font = Font(bold=True, size=16, color=PALETA["tinta"][1:])
    ws["A2"] = f"Período de {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(size=11, color=PALETA["marca_esc"][1:])

    identificacao = [
        ("Titular", (usuario.get_full_name() or usuario.username or "").strip() or "—"),
        ("Conteúdo", _rotulo_escopo(escopo)),
        ("Emitido em", timezone.localtime().strftime("%d/%m/%Y às %H:%M")),
    ]
    linha = 4
    for rotulo, valor in identificacao:
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=True)
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    linha += 1
    titulo = ws.cell(row=linha, column=1, value="INDICADORES DO PERÍODO")
    titulo.font = Font(bold=True, size=9, color=PALETA["marca_esc"][1:])
    linha += 1

    borda_indicador = Border(top=Side(style="thin", color=PALETA["linha"][1:]))
    for rotulo, valor, formato in indicadores:
        celula_rotulo = ws.cell(row=linha, column=1, value=rotulo)
        celula_rotulo.font = Font(bold=True)
        celula_rotulo.border = borda_indicador
        celula_valor = ws.cell(row=linha, column=2, value=valor)
        celula_valor.border = borda_indicador
        celula_valor.number_format = formato
        celula_valor.alignment = Alignment(horizontal="right")
        linha += 1

    if escopo in ("geral", "completo"):
        nota = (
            "Os valores consolidam lançamentos de caixa e faturas de cartão, sem duplicar "
            "as despesas individuais que compõem cada fatura."
        )
    else:
        nota = (
            "Patrimônio e resultado refletem a posição da custódia na data de emissão, "
            "avaliada pela cotação mais recente de cada ativo."
        )
    ws.cell(row=linha + 1, column=1, value=nota).font = Font(
        size=8, italic=True, color=PALETA["tinta_sec"][1:]
    )


def _dados_do_relatorio(usuario, data_inicio: date, data_fim: date, escopo: str) -> dict:
    """Reúne, em uma única passagem, tudo que os três formatos de saída consomem.

    Args:
        usuario (User): Usuário solicitante.
        data_inicio (date): Início do período.
        data_fim (date): Fim do período.
        escopo (str): Escopo já normalizado.

    Returns:
        dict: Conjuntos de dados e totais agregados do período.
    """
    inclui_geral = escopo in ("geral", "completo")
    inclui_investimentos = escopo in ("investimentos", "completo")

    comparativo = get_comparativo_mensal_data(usuario, data_inicio, data_fim) if inclui_geral else []
    categorias = get_despesas_por_categoria(usuario, data_inicio, data_fim) if inclui_geral else []
    movimentacoes = list(get_movimentacoes(usuario, data_inicio, data_fim)) if inclui_geral else []

    investimentos = (
        list(get_investimentos(usuario, data_inicio, data_fim)) if inclui_investimentos else []
    )
    transacoes = (
        list(get_transacoes_investimento(usuario, data_inicio, data_fim))
        if inclui_investimentos
        else []
    )
    proventos = (
        list(get_proventos_data(usuario, data_inicio, data_fim)) if inclui_investimentos else []
    )
    alocacao = get_alocacao_data(usuario, data_fim) if inclui_investimentos else []

    receitas = sum((_dec(item["receitas"]) for item in comparativo), Decimal("0.00"))
    despesas = sum((_dec(item["despesas"]) for item in comparativo), Decimal("0.00"))
    patrimonio = sum((_dec(item["valor"]) for item in alocacao), Decimal("0.00"))
    investido = sum((ativo.valor_investido for ativo in investimentos), Decimal("0.00"))
    mercado = sum((ativo.valor_total_atual for ativo in investimentos), Decimal("0.00"))

    return {
        "inclui_geral": inclui_geral,
        "inclui_investimentos": inclui_investimentos,
        "comparativo": comparativo,
        "categorias": categorias,
        "movimentacoes": movimentacoes,
        "investimentos": investimentos,
        "transacoes": transacoes,
        "proventos": proventos,
        "alocacao": alocacao,
        "receitas": receitas,
        "despesas": despesas,
        "resultado": receitas - despesas,
        "patrimonio": patrimonio,
        "investido": investido,
        "mercado": mercado,
        "resultado_carteira": mercado - investido,
        "proventos_total": sum((_dec(item["total"]) for item in proventos), Decimal("0.00")),
    }


def _indicadores(dados: dict, escopo: str) -> list:
    """Monta os pares rótulo/valor exibidos na aba Resumo.

    Args:
        dados (dict): Saída de `_dados_do_relatorio`.
        escopo (str): Escopo do relatório.

    Returns:
        list[tuple]: Trios (rótulo, valor, formato numérico) na ordem de exibição.
    """
    contagem = "#,##0"
    indicadores = []
    if dados["inclui_geral"]:
        indicadores += [
            ("Receitas do período", float(dados["receitas"]), FORMATO_MOEDA),
            ("Despesas do período", float(dados["despesas"]), FORMATO_MOEDA),
            ("Resultado do período", float(dados["resultado"]), FORMATO_MOEDA),
            ("Lançamentos no período", len(dados["movimentacoes"]), contagem),
        ]
    if dados["inclui_investimentos"]:
        indicadores += [
            ("Patrimônio investido", float(dados["patrimonio"]), FORMATO_MOEDA),
            ("Total investido (custo)", float(dados["investido"]), FORMATO_MOEDA),
            ("Resultado da carteira", float(dados["resultado_carteira"]), FORMATO_MOEDA),
            ("Proventos no período", float(dados["proventos_total"]), FORMATO_MOEDA),
            ("Ativos em carteira", len(dados["investimentos"]), contagem),
        ]
    return indicadores


def _linhas_movimentacoes(movimentacoes) -> tuple:
    """Prepara as linhas do extrato para planilha, com despesas negativas.

    Args:
        movimentacoes (Sequence[Conta]): Lançamentos do período.

    Returns:
        tuple: (linhas, resultado líquido) — o líquido é a soma da coluna de valor.
    """
    linhas = []
    liquido = Decimal("0.00")
    for movimentacao in movimentacoes:
        receita = movimentacao.tipo == Conta.TIPO_RECEITA
        valor = _dec(movimentacao.valor) if receita else -_dec(movimentacao.valor)
        liquido += valor
        linhas.append(
            [
                movimentacao.data_prevista,
                "Receita" if receita else "Despesa",
                movimentacao.descricao,
                movimentacao.categoria.nome if movimentacao.categoria else "Sem categoria",
                float(valor),
                "Realizada" if movimentacao.transacao_realizada else "Pendente",
            ]
        )
    return linhas, liquido


def _linhas_carteira(investimentos, total_mercado: Decimal) -> list:
    """Prepara as linhas da carteira, incluindo meta, valor ideal e sugestão.

    Args:
        investimentos (Sequence[Ativo]): Ativos do usuário.
        total_mercado (Decimal): Valor de mercado somado da carteira.

    Returns:
        list[list]: Linhas prontas para a planilha.
    """
    linhas = []
    for ativo in investimentos:
        valor_mercado = ativo.valor_total_atual
        valor_investido = ativo.valor_investido
        meta = _dec(ativo.meta_porcentagem)
        valor_ideal = (meta / 100) * total_mercado if total_mercado > 0 else Decimal("0.00")
        linhas.append(
            [
                ativo.ticker,
                ativo.nome or "",
                ativo.subcategoria.categoria.classe.nome if ativo.subcategoria else "Sem classe",
                ativo.subcategoria.categoria.nome if ativo.subcategoria else "Sem categoria",
                float(ativo.quantidade),
                float(ativo.preco_medio),
                float(valor_investido),
                float(valor_mercado),
                float(meta / 100),
                float(valor_ideal),
                float(valor_ideal - valor_mercado),
                float(valor_mercado - valor_investido),
            ]
        )
    return linhas


def gerar_excel(usuario, data_inicio: date, data_fim: date, escopo: str = "completo") -> bytes:
    """Gera a planilha Excel (.xlsx) do relatório financeiro.

    A primeira aba traz a identificação e os indicadores do período; as demais
    trazem uma tabela por assunto, todas com cabeçalho congelado, filtro
    automático, formatos numéricos nativos (somáveis no Excel) e linha de total.

    Args:
        usuario (User): Instância do usuário Django solicitante.
        data_inicio (date): Data de início para o filtro do relatório.
        data_fim (date): Data final para o filtro do relatório.
        escopo (str, optional): Escopo do relatório ('geral', 'investimentos', 'completo'). Defaults to "completo".

    Returns:
        bytes: O conteúdo em bytes da planilha gerada em formato openxml (.xlsx).
    """
    if escopo not in ("geral", "investimentos", "completo"):
        escopo = "completo"
    dados = _dados_do_relatorio(usuario, data_inicio, data_fim, escopo)

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = (
        f"Relatório Financeiro {data_inicio.strftime('%d-%m-%Y')} a {data_fim.strftime('%d-%m-%Y')}"
    )
    wb.properties.creator = "FreeCash"
    wb.properties.subject = _rotulo_escopo(escopo)

    cor_geral = PALETA["marca"]
    cor_investimentos = SERIE_CORES[1]

    if dados["inclui_geral"]:
        linhas, liquido = _linhas_movimentacoes(dados["movimentacoes"])
        _montar_aba(
            wb,
            "Movimentações",
            ["Data", "Tipo", "Descrição", "Categoria", "Valor (R$)", "Situação"],
            linhas,
            [FORMATO_DATA, None, None, None, FORMATO_MOEDA, None],
            cor_geral,
            linha_total=[
                f"{len(linhas)} lançamentos", "", "", "", float(liquido), ""
            ] if linhas else None,
        )

        linhas_mes = [
            [
                date(int(item["periodo"][:4]), int(item["periodo"][5:7]), 1),
                float(_dec(item["receitas"])),
                float(_dec(item["despesas"])),
                float(_dec(item["saldo"])),
            ]
            for item in dados["comparativo"]
        ]
        _montar_aba(
            wb,
            "Resumo Mensal",
            ["Mês", "Receitas (R$)", "Despesas (R$)", "Resultado (R$)"],
            linhas_mes,
            [FORMATO_MES, FORMATO_MOEDA, FORMATO_MOEDA, FORMATO_MOEDA],
            cor_geral,
            linha_total=[
                "Total do período",
                float(dados["receitas"]),
                float(dados["despesas"]),
                float(dados["resultado"]),
            ] if linhas_mes else None,
        )

        linhas_categoria = [
            [item["categoria"], float(_dec(item["total"])), float(_dec(item["percentual"]) / 100)]
            for item in dados["categorias"]
        ]
        _montar_aba(
            wb,
            "Despesas por Categoria",
            ["Categoria", "Total (R$)", "% das despesas"],
            linhas_categoria,
            [None, FORMATO_MOEDA, FORMATO_PERCENTUAL],
            cor_geral,
            linha_total=["Total", float(dados["despesas"]), 1.0] if linhas_categoria else None,
        )

    if dados["inclui_investimentos"]:
        linhas_carteira = _linhas_carteira(dados["investimentos"], dados["mercado"])
        _montar_aba(
            wb,
            "Carteira",
            [
                "Ticker", "Nome", "Classe", "Categoria", "Quantidade", "Preço médio (R$)",
                "Investido (R$)", "Mercado (R$)", "Meta", "Valor ideal (R$)",
                "Sugestão (R$)", "Resultado (R$)",
            ],
            linhas_carteira,
            [
                None, None, None, None, FORMATO_QUANTIDADE, FORMATO_MOEDA, FORMATO_MOEDA,
                FORMATO_MOEDA, FORMATO_PERCENTUAL, FORMATO_MOEDA, FORMATO_MOEDA, FORMATO_MOEDA,
            ],
            cor_investimentos,
            # Os totais de "ideal" e "sugestão" são somados linha a linha: assumir
            # que fecham com o valor de mercado só valeria se as metas somassem 100%.
            linha_total=[
                "Total", "", "", "", "", "",
                float(dados["investido"]),
                float(dados["mercado"]),
                sum(linha[8] for linha in linhas_carteira),
                sum(linha[9] for linha in linhas_carteira),
                sum(linha[10] for linha in linhas_carteira),
                float(dados["resultado_carteira"]),
            ] if linhas_carteira else None,
        )

        linhas_alocacao = [
            [item["classe"], float(_dec(item["valor"])), float(_dec(item["percentual"]) / 100)]
            for item in dados["alocacao"]
        ]
        _montar_aba(
            wb,
            "Alocação",
            ["Classe", "Valor (R$)", "% da carteira"],
            linhas_alocacao,
            [None, FORMATO_MOEDA, FORMATO_PERCENTUAL],
            cor_investimentos,
            linha_total=["Total", float(dados["patrimonio"]), 1.0] if linhas_alocacao else None,
        )

        linhas_proventos = [
            [item["ativo__ticker"], float(_dec(item["total"]))] for item in dados["proventos"]
        ]
        _montar_aba(
            wb,
            "Proventos",
            ["Ticker", "Total recebido (R$)"],
            linhas_proventos,
            [None, FORMATO_MOEDA],
            cor_investimentos,
            linha_total=["Total", float(dados["proventos_total"])] if linhas_proventos else None,
        )

        linhas_transacoes = [
            [
                transacao.data,
                transacao.ativo.ticker,
                transacao.get_tipo_display(),
                float(transacao.quantidade),
                float(transacao.preco_unitario),
                float(transacao.taxas),
                float(transacao.valor_total),
            ]
            for transacao in dados["transacoes"]
        ]
        _montar_aba(
            wb,
            "Transações Invest.",
            ["Data", "Ticker", "Operação", "Quantidade", "Preço (R$)", "Taxas (R$)", "Total (R$)"],
            linhas_transacoes,
            [
                FORMATO_DATA, None, None, FORMATO_QUANTIDADE, FORMATO_MOEDA,
                FORMATO_MOEDA, FORMATO_MOEDA,
            ],
            cor_investimentos,
            linha_total=[
                f"{len(linhas_transacoes)} ordens", "", "", "", "", "",
                float(sum(_dec(t.valor_total) for t in dados["transacoes"])),
            ] if linhas_transacoes else None,
        )

    _aba_resumo(wb, usuario, data_inicio, data_fim, escopo, _indicadores(dados, escopo))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def gerar_csv(usuario, data_inicio: date, data_fim: date, escopo: str = "completo") -> str:
    """Gera o relatório em CSV pronto para abrir no Excel/LibreOffice em pt-BR.

    Usa ponto e vírgula como separador, vírgula decimal e prefixo BOM — a
    combinação que faz o Excel brasileiro reconhecer colunas, números e acentos
    sem passar pelo assistente de importação.

    Args:
        usuario (User): Instância do usuário Django solicitante.
        data_inicio (date): Data de início para o filtro do relatório.
        data_fim (date): Data final para o filtro do relatório.
        escopo (str, optional): Escopo do relatório ('geral', 'investimentos', 'completo'). Defaults to "completo".

    Returns:
        str: Conteúdo do arquivo CSV, iniciado por BOM UTF-8.
    """
    if escopo not in ("geral", "investimentos", "completo"):
        escopo = "completo"
    dados = _dados_do_relatorio(usuario, data_inicio, data_fim, escopo)

    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";", lineterminator="\r\n")

    def secao(titulo: str, cabecalhos) -> None:
        """Abre uma nova seção com título e cabeçalho de colunas."""
        escritor.writerow([])
        escritor.writerow([titulo])
        escritor.writerow(list(cabecalhos))

    titular = (usuario.get_full_name() or usuario.username or "").strip() or "—"
    escritor.writerow(["Relatório Financeiro FreeCash"])
    escritor.writerow(["Titular", titular])
    escritor.writerow(
        [
            "Período",
            data_inicio.strftime("%d/%m/%Y"),
            data_fim.strftime("%d/%m/%Y"),
        ]
    )
    escritor.writerow(["Conteúdo", _rotulo_escopo(escopo)])
    escritor.writerow(["Emitido em", timezone.localtime().strftime("%d/%m/%Y %H:%M")])

    escritor.writerow([])
    escritor.writerow(["INDICADORES"])
    for rotulo, valor, formato in _indicadores(dados, escopo):
        if formato == FORMATO_MOEDA:
            escritor.writerow([rotulo, formatar_moeda(valor, simbolo=False)])
        else:
            escritor.writerow([rotulo, formatar_numero(valor, 0)])

    if dados["inclui_geral"]:
        secao(
            "MOVIMENTAÇÕES",
            ["Data", "Tipo", "Descrição", "Categoria", "Valor (R$)", "Situação"],
        )
        linhas, liquido = _linhas_movimentacoes(dados["movimentacoes"])
        for data_prevista, tipo, descricao, categoria, valor, situacao in linhas:
            escritor.writerow(
                [
                    data_prevista.strftime("%d/%m/%Y"),
                    tipo,
                    descricao,
                    categoria,
                    formatar_moeda(valor, simbolo=False),
                    situacao,
                ]
            )
        if linhas:
            escritor.writerow(
                [
                    f"{len(linhas)} lançamentos", "", "", "",
                    formatar_moeda(liquido, simbolo=False), "",
                ]
            )
        else:
            escritor.writerow(["Nenhuma movimentação no período"])

        secao(
            "RESUMO MENSAL",
            ["Mês", "Receitas (R$)", "Despesas (R$)", "Resultado (R$)"],
        )
        for item in dados["comparativo"]:
            escritor.writerow(
                [
                    formatar_mes_ano(item["periodo"]),
                    formatar_moeda(item["receitas"], simbolo=False),
                    formatar_moeda(item["despesas"], simbolo=False),
                    formatar_moeda(item["saldo"], simbolo=False),
                ]
            )
        escritor.writerow(
            [
                "Total do período",
                formatar_moeda(dados["receitas"], simbolo=False),
                formatar_moeda(dados["despesas"], simbolo=False),
                formatar_moeda(dados["resultado"], simbolo=False),
            ]
        )

        secao("DESPESAS POR CATEGORIA", ["Categoria", "Total (R$)", "% das despesas"])
        for item in dados["categorias"]:
            escritor.writerow(
                [
                    item["categoria"],
                    formatar_moeda(item["total"], simbolo=False),
                    formatar_percentual(item["percentual"]),
                ]
            )

    if dados["inclui_investimentos"]:
        secao(
            "CARTEIRA DE INVESTIMENTOS",
            [
                "Ticker", "Nome", "Classe", "Categoria", "Quantidade", "Preço médio (R$)",
                "Investido (R$)", "Mercado (R$)", "Meta", "Valor ideal (R$)",
                "Sugestão (R$)", "Resultado (R$)",
            ],
        )
        for linha in _linhas_carteira(dados["investimentos"], dados["mercado"]):
            escritor.writerow(
                [
                    linha[0], linha[1], linha[2], linha[3],
                    formatar_numero(linha[4]),
                    formatar_moeda(linha[5], simbolo=False),
                    formatar_moeda(linha[6], simbolo=False),
                    formatar_moeda(linha[7], simbolo=False),
                    formatar_percentual(_dec(linha[8]) * 100),
                    formatar_moeda(linha[9], simbolo=False),
                    formatar_moeda(linha[10], simbolo=False),
                    formatar_moeda(linha[11], simbolo=False),
                ]
            )
        if dados["investimentos"]:
            escritor.writerow(
                [
                    "Total", "", "", "", "", "",
                    formatar_moeda(dados["investido"], simbolo=False),
                    formatar_moeda(dados["mercado"], simbolo=False),
                    "", "", "",
                    formatar_moeda(dados["resultado_carteira"], simbolo=False),
                ]
            )
        else:
            escritor.writerow(["Nenhum ativo em carteira no período"])

        secao("ALOCAÇÃO POR CLASSE", ["Classe", "Valor (R$)", "% da carteira"])
        for item in dados["alocacao"]:
            escritor.writerow(
                [
                    item["classe"],
                    formatar_moeda(item["valor"], simbolo=False),
                    formatar_percentual(item["percentual"]),
                ]
            )

        secao("PROVENTOS RECEBIDOS", ["Ticker", "Total recebido (R$)"])
        for item in dados["proventos"]:
            escritor.writerow(
                [item["ativo__ticker"], formatar_moeda(item["total"], simbolo=False)]
            )
        if dados["proventos"]:
            escritor.writerow(
                ["Total", formatar_moeda(dados["proventos_total"], simbolo=False)]
            )

        secao(
            "TRANSAÇÕES DE INVESTIMENTO",
            ["Data", "Ticker", "Operação", "Quantidade", "Preço (R$)", "Taxas (R$)", "Total (R$)"],
        )
        for transacao in dados["transacoes"]:
            escritor.writerow(
                [
                    transacao.data.strftime("%d/%m/%Y"),
                    transacao.ativo.ticker,
                    transacao.get_tipo_display(),
                    formatar_numero(transacao.quantidade),
                    formatar_moeda(transacao.preco_unitario, simbolo=False),
                    formatar_moeda(transacao.taxas, simbolo=False),
                    formatar_moeda(transacao.valor_total, simbolo=False),
                ]
            )

    # BOM para o Excel reconhecer UTF-8 e exibir os acentos corretamente
    return "﻿" + buffer.getvalue()


# =============================================================================
# PDF — ESTILOS E COMPONENTES REUTILIZÁVEIS
# =============================================================================


def _estilos() -> dict:
    """Monta os estilos de parágrafo usados no relatório PDF.

    Returns:
        dict: Mapa de nome do estilo para ParagraphStyle configurado.
    """
    base = getSampleStyleSheet()
    estilos = {
        "capa_titulo": ParagraphStyle(
            "CapaTitulo", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=26, leading=29, textColor=cor("tinta"), spaceAfter=3,
        ),
        "capa_periodo": ParagraphStyle(
            "CapaPeriodo", parent=base["Normal"], fontName="Helvetica",
            fontSize=12.5, leading=16, textColor=cor("marca_esc"), spaceAfter=10,
        ),
        "capa_meta": ParagraphStyle(
            "CapaMeta", parent=base["Normal"], fontName="Helvetica",
            fontSize=8.5, leading=13, textColor=cor("tinta_sec"),
        ),
        "secao": ParagraphStyle(
            "Secao", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=12, leading=15, textColor=cor("marca_esc"),
        ),
        "secao_nota": ParagraphStyle(
            "SecaoNota", parent=base["Normal"], fontName="Helvetica",
            fontSize=8, leading=11, textColor=cor("tinta_sec"),
        ),
        "corpo": ParagraphStyle(
            "Corpo", parent=base["Normal"], fontName="Helvetica",
            fontSize=9, leading=13, textColor=cor("tinta"),
        ),
        "nota": ParagraphStyle(
            "Nota", parent=base["Normal"], fontName="Helvetica",
            fontSize=7.5, leading=10, textColor=cor("tinta_sec"),
        ),
        "vazio": ParagraphStyle(
            "Vazio", parent=base["Normal"], fontName="Helvetica-Oblique",
            fontSize=9, leading=13, textColor=cor("tinta_sec"),
        ),
        "kpi_rotulo": ParagraphStyle(
            "KpiRotulo", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=6.5, leading=9, alignment=TA_CENTER, textColor=cor("tinta_sec"),
        ),
        "kpi_rotulo_esq": ParagraphStyle(
            "KpiRotuloEsq", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=7, leading=10, alignment=TA_LEFT, textColor=cor("marca_esc"),
        ),
    }
    for chave in ("positivo", "negativo", "tinta", "marca", "marca_esc"):
        estilos[f"kpi_{chave}"] = ParagraphStyle(
            f"KpiValor{chave}", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=12.5, leading=15, alignment=TA_CENTER, textColor=cor(chave),
        )
    return estilos


#: Cache de estilos de célula, indexado por (corpo da fonte, cor, alinhamento, peso).
_CACHE_ESTILO_CELULA = {}


def _estilo_celula(
    fonte: float, chave_cor: str = "tinta", direita: bool = False, negrito: bool = False
) -> ParagraphStyle:
    """Obtém (com cache) o estilo de célula para o corpo de fonte e cor informados.

    Args:
        fonte (float): Corpo da fonte da tabela.
        chave_cor (str, optional): Cor da paleta aplicada ao texto. Defaults to "tinta".
        direita (bool, optional): Alinha o conteúdo à direita. Defaults to False.
        negrito (bool, optional): Usa a variante em negrito. Defaults to False.

    Returns:
        ParagraphStyle: Estilo pronto para envolver o conteúdo da célula.
    """
    chave = (round(float(fonte), 2), chave_cor, bool(direita), bool(negrito))
    if chave not in _CACHE_ESTILO_CELULA:
        _CACHE_ESTILO_CELULA[chave] = ParagraphStyle(
            "cel-%s-%s-%d-%d" % (chave[0], chave_cor, direita, negrito),
            fontName="Helvetica-Bold" if negrito else "Helvetica",
            fontSize=fonte,
            leading=fonte * 1.22,
            textColor=cor(chave_cor),
            alignment=TA_RIGHT if direita else TA_LEFT,
        )
    return _CACHE_ESTILO_CELULA[chave]


def _texto(conteudo, estilo) -> Paragraph:
    """Cria um parágrafo escapando caracteres reservados de marcação.

    Args:
        conteudo: Texto de origem (aceita None).
        estilo (ParagraphStyle): Estilo a ser aplicado.

    Returns:
        Paragraph: Flowable pronto para uso em células de tabela ou no corpo.
    """
    return Paragraph(escape_xml(str(conteudo or "")), estilo)


def _valor_colorido(valor, fonte: float = 7.5, negrito: bool = False) -> Paragraph:
    """Renderiza um valor monetário alinhado à direita e colorido pelo sinal.

    Valores nulos ficam em cor neutra para não sugerir ganho onde não houve.

    Args:
        valor: Valor monetário.
        fonte (float, optional): Corpo da fonte da tabela. Defaults to 7.5.
        negrito (bool, optional): Usa negrito (linhas de totalização). Defaults to False.

    Returns:
        Paragraph: Célula em verde (> 0), vermelho (< 0) ou grafite (= 0).
    """
    valor_decimal = _dec(valor)
    if valor_decimal > 0:
        chave = "positivo"
    elif valor_decimal < 0:
        chave = "negativo"
    else:
        chave = "tinta"
    return Paragraph(
        formatar_moeda(valor_decimal),
        _estilo_celula(fonte, chave, direita=True, negrito=negrito),
    )


def _larguras(largura_total: float, fracoes) -> list:
    """Converte proporções de coluna em larguras absolutas normalizadas.

    Args:
        largura_total (float): Largura útil do frame (doc.width).
        fracoes (Sequence[float]): Proporções relativas de cada coluna.

    Returns:
        list[float]: Larguras em pontos que somam exatamente a largura útil.
    """
    soma = float(sum(fracoes)) or 1.0
    return [largura_total * float(f) / soma for f in fracoes]


def _tabela(
    dados,
    larguras,
    *,
    numericas=(),
    total: bool = False,
    fonte: float = 7.5,
) -> Table:
    """Cria uma tabela com o visual institucional único do relatório.

    Centraliza cabeçalho colorido, repetição de cabeçalho entre páginas, zebra,
    espaçamento interno e alinhamento das colunas numéricas. Toda célula de dados
    é convertida em parágrafo, o que garante quebra de linha em textos longos
    (nada é truncado) e alinhamento vertical uniforme entre as colunas.

    Args:
        dados (list[list]): Matriz de células; a primeira linha é o cabeçalho.
        larguras (list[float]): Larguras das colunas em pontos.
        numericas (Sequence[int], optional): Índices de colunas alinhadas à direita.
        total (bool, optional): Destaca a última linha como totalizador. Defaults to False.
        fonte (float, optional): Corpo da fonte das células. Defaults to 7.5.

    Returns:
        Table: Flowable estilizado, com cabeçalho repetido em quebras de página.
    """
    ultima_linha = len(dados) - 1
    corpo = [list(dados[0])]
    for indice, linha in enumerate(dados[1:], start=1):
        negrito = total and indice == ultima_linha
        corpo.append(
            [
                celula
                if hasattr(celula, "wrapOn")
                else Paragraph(
                    escape_xml(str(celula if celula is not None else "")),
                    _estilo_celula(
                        fonte, direita=coluna in numericas, negrito=negrito
                    ),
                )
                for coluna, celula in enumerate(linha)
            ]
        )

    tabela = Table(corpo, colWidths=larguras, repeatRows=1, hAlign="LEFT")
    comandos = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), fonte),
        ("TEXTCOLOR", (0, 1), (-1, -1), cor("tinta")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, cor("linha")),
        ("BACKGROUND", (0, 0), (-1, 0), cor("marca_esc")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), max(fonte - 0.7, 6)),
        ("TOPPADDING", (0, 0), (-1, 0), 5.5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, cor("zebra")]),
    ]
    for coluna in numericas:
        comandos.append(("ALIGN", (coluna, 0), (coluna, -1), "RIGHT"))
    if total:
        comandos += [
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), cor("destaque")),
            ("LINEABOVE", (0, -1), (-1, -1), 0.9, cor("marca")),
            ("TOPPADDING", (0, -1), (-1, -1), 5),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 5),
        ]
    tabela.setStyle(TableStyle(comandos))
    return tabela


def _titulo_secao(titulo: str, largura: float, estilos: dict, nota: str = "") -> list:
    """Monta o cabeçalho de uma seção: barra de acento, título e régua da marca.

    Args:
        titulo (str): Nome da seção.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.
        nota (str, optional): Linha explicativa exibida abaixo do título.

    Returns:
        list: Sequência de flowables a serem estendidos na história do documento.
    """
    cabecalho = Table(
        [[_texto(titulo.upper(), estilos["secao"])]], colWidths=[largura], hAlign="LEFT"
    )
    cabecalho.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LINEBEFORE", (0, 0), (0, 0), 2.5, cor("marca")),
                ("LINEBELOW", (0, 0), (-1, 0), 0.75, cor("linha")),
            ]
        )
    )
    elementos = [cabecalho, Spacer(1, 3.5 * mm)]
    if nota:
        elementos.insert(1, Spacer(1, 1.5 * mm))
        elementos.insert(2, _texto(nota, estilos["secao_nota"]))
    return elementos


def _cartoes_kpi(itens, largura: float, estilos: dict) -> Table:
    """Monta a faixa de indicadores da capa em formato de cartões.

    Args:
        itens (Sequence[tuple]): Tuplas (rótulo, valor formatado, chave de cor).
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        Table: Grade de cartões com acento colorido no topo de cada indicador.
    """
    rotulos = [_texto(rotulo.upper(), estilos["kpi_rotulo"]) for rotulo, _, _ in itens]
    valores = [
        Paragraph(valor, estilos[f"kpi_{chave_cor}"]) for _, valor, chave_cor in itens
    ]
    tabela = Table(
        [rotulos, valores],
        colWidths=_larguras(largura, [1] * len(itens)),
        hAlign="LEFT",
    )
    comandos = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
        ("TOPPADDING", (0, 1), (-1, 1), 0),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (0, 0), (-1, -1), cor("zebra")),
        ("BOX", (0, 0), (-1, -1), 0.25, cor("linha")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.white),
    ]
    for coluna, (_, _, chave_cor) in enumerate(itens):
        comandos.append(("LINEABOVE", (coluna, 0), (coluna, 0), 2.2, cor(chave_cor)))
    tabela.setStyle(TableStyle(comandos))
    return tabela


# =============================================================================
# PDF — GRÁFICOS
# =============================================================================


def render_grafico_alocacao(alocacao_dados, largura: float = 440) -> Drawing:
    """Gera o gráfico de rosca da alocação da carteira por classe de ativo.

    Os rótulos ficam apenas na legenda (com valor e percentual), evitando a
    sobreposição de texto sobre as fatias, e todas as fatias recebem cor
    explícita independentemente da quantidade de classes.

    Args:
        alocacao_dados (list[dict]): Itens com 'classe', 'valor' e 'percentual'.
        largura (float, optional): Largura útil disponível em pontos. Defaults to 440.

    Returns:
        Drawing: Desenho vetorial dimensionado para o frame, ou None se não houver dados.
    """
    if not alocacao_dados:
        return None

    altura = 200.0
    desenho = Drawing(largura, altura)
    total = sum((_dec(item["valor"]) for item in alocacao_dados), Decimal("0.00"))

    rosca = Doughnut()
    rosca.x = 6
    rosca.y = 22
    rosca.width = 148
    rosca.height = 148
    rosca.data = [float(_dec(item["valor"])) for item in alocacao_dados]
    rosca.labels = None
    rosca.innerRadiusFraction = 0.58
    rosca.slices.strokeColor = colors.white
    rosca.slices.strokeWidth = 1.2
    for indice in range(len(rosca.data)):
        rosca.slices[indice].fillColor = cor_serie(indice)
    desenho.add(rosca)

    centro_x = rosca.x + rosca.width / 2
    centro_y = rosca.y + rosca.height / 2
    desenho.add(
        String(
            centro_x, centro_y + 5, "PATRIMÔNIO",
            fontName="Helvetica-Bold", fontSize=6.5,
            fillColor=cor("tinta_sec"), textAnchor="middle",
        )
    )
    desenho.add(
        String(
            centro_x, centro_y - 9, formatar_moeda(total),
            fontName="Helvetica-Bold", fontSize=9.5,
            fillColor=cor("tinta"), textAnchor="middle",
        )
    )

    passo_legenda = 13
    legenda = Legend()
    legenda.x = rosca.x + rosca.width + 20
    # Centraliza verticalmente a legenda em relação à rosca
    legenda.y = min(
        altura - 12,
        centro_y + len(alocacao_dados) * passo_legenda / 2,
    )
    legenda.boxAnchor = "nw"
    legenda.alignment = "right"  # desenha o quadrado de cor antes do texto
    legenda.dx = 7
    legenda.dy = 7
    legenda.dxTextSpace = 6
    legenda.deltay = passo_legenda
    legenda.fontName = "Helvetica"
    legenda.fontSize = 7.5
    legenda.columnMaximum = 12
    legenda.strokeWidth = 0
    legenda.strokeColor = None
    legenda.colorNamePairs = [
        (
            cor_serie(indice),
            f"{item['classe']}   {formatar_moeda(item['valor'])}  ({formatar_percentual(item['percentual'])})",
        )
        for indice, item in enumerate(alocacao_dados)
    ]
    desenho.add(legenda)
    return desenho


def render_grafico_evolucao_mensal(
    comparativo, largura: float = 440, meses: int = MESES_NO_GRAFICO
) -> Drawing:
    """Gera o gráfico de barras de receitas x despesas por mês.

    Args:
        comparativo (list[dict]): Saída de `get_comparativo_mensal_data`.
        largura (float, optional): Largura útil disponível em pontos. Defaults to 440.
        meses (int, optional): Quantidade de períodos mais recentes exibidos. Defaults to 12.

    Returns:
        Drawing: Desenho vetorial do gráfico, ou None se não houver dados.
    """
    if not comparativo:
        return None

    dados = comparativo[-meses:] if meses else comparativo
    altura = 215.0
    desenho = Drawing(largura, altura)

    receitas = [float(_dec(item["receitas"])) for item in dados]
    despesas = [float(_dec(item["despesas"])) for item in dados]

    grafico = VerticalBarChart()
    grafico.x = 48
    grafico.y = 50
    grafico.width = max(largura - grafico.x - 12, 60)
    grafico.height = altura - grafico.y - 32
    grafico.data = [receitas, despesas]
    grafico.barSpacing = 1
    grafico.groupSpacing = 7
    grafico.bars.strokeWidth = 0
    grafico.bars[0].fillColor = cor("marca")
    grafico.bars[1].fillColor = cor("negativo")

    grafico.categoryAxis.categoryNames = [formatar_mes_ano(item["periodo"]) for item in dados]
    grafico.categoryAxis.labels.fontName = "Helvetica"
    grafico.categoryAxis.labels.fontSize = 6.5
    grafico.categoryAxis.labels.angle = 40
    grafico.categoryAxis.labels.boxAnchor = "e"
    grafico.categoryAxis.labels.dy = -3
    grafico.categoryAxis.strokeColor = cor("linha")

    grafico.valueAxis.valueMin = 0
    grafico.valueAxis.valueMax = _teto_eixo(max(receitas + despesas + [1.0]))
    grafico.valueAxis.valueStep = grafico.valueAxis.valueMax / 4.0
    grafico.valueAxis.labels.fontName = "Helvetica"
    grafico.valueAxis.labels.fontSize = 6.5
    grafico.valueAxis.labelTextFormat = lambda valor: formatar_numero(valor, 0)
    grafico.valueAxis.strokeColor = cor("linha")
    grafico.valueAxis.visibleGrid = 1
    grafico.valueAxis.gridStrokeColor = cor("linha")
    grafico.valueAxis.gridStrokeWidth = 0.25
    desenho.add(grafico)

    desenho.add(
        String(
            0, altura - 11, "Receitas x despesas por mês",
            fontName="Helvetica-Bold", fontSize=9.5, fillColor=cor("tinta"),
        )
    )
    desenho.add(
        String(
            0, altura - 23, "Valores em reais (R$)",
            fontName="Helvetica", fontSize=7, fillColor=cor("tinta_sec"),
        )
    )

    legenda = Legend()
    legenda.x = grafico.x
    legenda.y = 14
    legenda.boxAnchor = "nw"
    legenda.alignment = "right"  # desenha o quadrado de cor antes do texto
    legenda.dx = 7
    legenda.dy = 7
    legenda.dxTextSpace = 5
    legenda.deltax = 82
    legenda.columnMaximum = 1
    legenda.fontName = "Helvetica"
    legenda.fontSize = 7.5
    legenda.strokeWidth = 0
    legenda.strokeColor = None
    legenda.colorNamePairs = [
        (cor("marca"), "Receitas"),
        (cor("negativo"), "Despesas"),
    ]
    desenho.add(legenda)
    return desenho


def render_grafico_waterfall(total_receitas, total_despesas, largura: float = 440) -> Drawing:
    """Gera o gráfico em cascata da composição do resultado do período.

    Mostra receitas, o abatimento das despesas e o resultado final, incluindo o
    caso de resultado negativo (barra abaixo da linha do zero).

    Args:
        total_receitas: Soma das receitas do período.
        total_despesas: Soma das despesas do período.
        largura (float, optional): Largura útil disponível em pontos. Defaults to 440.

    Returns:
        Drawing: Desenho vetorial do gráfico em cascata.
    """
    receitas = _dec(total_receitas)
    despesas = _dec(total_despesas)
    resultado = receitas - despesas

    altura = 178.0
    desenho = Drawing(largura, altura)
    base = 42.0
    area = 108.0

    topo = max(float(receitas), float(resultado), 0.0)
    piso = min(0.0, float(resultado))
    amplitude = (topo - piso) or 1.0

    def posicao_y(valor) -> float:
        return base + (float(valor) - piso) / amplitude * area

    etapas = [
        ("Receitas", posicao_y(0), posicao_y(receitas), "positivo", receitas),
        ("Despesas", posicao_y(resultado), posicao_y(receitas), "negativo", -despesas),
        (
            "Resultado",
            min(posicao_y(0), posicao_y(resultado)),
            max(posicao_y(0), posicao_y(resultado)),
            "marca" if resultado >= 0 else "negativo",
            resultado,
        ),
    ]

    vao = largura / len(etapas)
    largura_barra = min(76.0, vao * 0.42)
    y_zero = posicao_y(0)

    desenho.add(
        String(
            0, altura - 11, "Composição do resultado do período",
            fontName="Helvetica-Bold", fontSize=9.5, fillColor=cor("tinta"),
        )
    )
    desenho.add(
        Line(
            0, y_zero, largura, y_zero,
            strokeColor=cor("linha"), strokeWidth=0.75,
        )
    )

    for indice, (rotulo, y_inicio, y_fim, chave_cor, valor) in enumerate(etapas):
        centro = vao * (indice + 0.5)
        x_barra = centro - largura_barra / 2
        altura_barra = max(y_fim - y_inicio, 1.2)
        desenho.add(
            Rect(
                x_barra, y_inicio, largura_barra, altura_barra,
                fillColor=cor(chave_cor), strokeColor=None,
            )
        )
        desenho.add(
            String(
                centro, y_fim + 6, formatar_moeda(valor),
                fontName="Helvetica-Bold", fontSize=8,
                fillColor=cor(chave_cor), textAnchor="middle",
            )
        )
        desenho.add(
            String(
                centro, base - 16, rotulo.upper(),
                fontName="Helvetica-Bold", fontSize=7,
                fillColor=cor("tinta_sec"), textAnchor="middle",
            )
        )
        if indice < len(etapas) - 1:
            y_conector = y_fim if indice == 0 else min(etapas[indice][1], etapas[indice][2])
            desenho.add(
                Line(
                    x_barra + largura_barra, y_conector,
                    vao * (indice + 1.5) - largura_barra / 2, y_conector,
                    strokeColor=cor("tinta_sec"), strokeWidth=0.5,
                    strokeDashArray=[2, 2],
                )
            )

    return desenho


# =============================================================================
# PDF — MOLDURA DA PÁGINA
# =============================================================================


def _desenhar_wordmark(canv, x: float, y: float, sobre_faixa: bool = False) -> None:
    """Desenha a marca "FreeCash" vetorialmente, sem depender de arquivo de imagem.

    Args:
        canv (Canvas): Canvas ativo do reportlab.
        x (float): Coordenada horizontal da base da marca.
        y (float): Linha de base do texto da marca.
        sobre_faixa (bool, optional): Inverte as cores para uso sobre a faixa azul.
    """
    lado = 4.6 * mm
    canv.saveState()
    canv.setFillColor(colors.white if sobre_faixa else cor("marca"))
    canv.roundRect(x, y - 0.9 * mm, lado, lado, 1.1 * mm, stroke=0, fill=1)
    canv.setFillColor(cor("marca") if sobre_faixa else colors.white)
    canv.setFont("Helvetica-Bold", 8)
    canv.drawCentredString(x + lado / 2, y + 0.5 * mm, "F")
    canv.setFillColor(colors.white if sobre_faixa else cor("tinta"))
    canv.setFont("Helvetica-Bold", 11)
    canv.drawString(x + lado + 2.2 * mm, y, "FreeCash")
    canv.restoreState()


def _construir_moldura(periodo_texto: str, emissao_texto: str):
    """Cria os callbacks de página da capa e das páginas internas.

    Args:
        periodo_texto (str): Período do relatório exibido no cabeçalho corrido.
        emissao_texto (str): Linha de emissão exibida no rodapé da capa.

    Returns:
        tuple: Par de callbacks (capa, páginas internas) para `doc.build`.
    """

    def _capa(canv, doc):
        largura, altura = A4
        canv.saveState()
        canv.setFillColor(cor("marca"))
        canv.rect(0, altura - MARGEM_TOPO, largura, MARGEM_TOPO, stroke=0, fill=1)
        _desenhar_wordmark(canv, MARGEM_LATERAL, altura - MARGEM_TOPO + 8 * mm, sobre_faixa=True)
        canv.setFont("Helvetica", 8)
        canv.setFillColor(colors.white)
        canv.drawRightString(
            largura - MARGEM_LATERAL, altura - MARGEM_TOPO + 8.5 * mm, "Relatório Financeiro"
        )
        canv.setFont("Helvetica", 7)
        canv.setFillColor(cor("tinta_sec"))
        canv.drawString(MARGEM_LATERAL, MARGEM_BASE - 8 * mm, emissao_texto)
        canv.restoreState()

    def _interna(canv, doc):
        largura, altura = A4
        linha_base = altura - MARGEM_TOPO + 6 * mm
        canv.saveState()
        _desenhar_wordmark(canv, MARGEM_LATERAL, linha_base + 1.5 * mm)
        canv.setFont("Helvetica", 7.5)
        canv.setFillColor(cor("tinta_sec"))
        canv.drawRightString(largura - MARGEM_LATERAL, linha_base + 2 * mm, periodo_texto)
        canv.setStrokeColor(cor("marca"))
        canv.setLineWidth(0.75)
        canv.line(
            MARGEM_LATERAL, linha_base - 1 * mm, largura - MARGEM_LATERAL, linha_base - 1 * mm
        )
        canv.restoreState()

    return _capa, _interna


class CanvasNumerado(pdfcanvas.Canvas):
    """Canvas em duas passagens que imprime "Página X de Y" no rodapé.

    O total de páginas só é conhecido no fim da geração, então os estados de
    página são acumulados e reescritos em `save()`. A capa não recebe rodapé
    numerado para não competir com a identidade do documento.
    """

    def __init__(self, *args, **kwargs):
        kwargs.setdefault("lang", "pt-BR")
        super().__init__(*args, **kwargs)
        self._paginas_pendentes = []

    def showPage(self):
        """Guarda o estado da página corrente em vez de finalizá-la imediatamente."""
        self._paginas_pendentes.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """Reescreve cada página acumulada com o rodapé numerado e finaliza o arquivo."""
        total = len(self._paginas_pendentes)
        for indice, estado in enumerate(self._paginas_pendentes, start=1):
            self.__dict__.update(estado)
            if indice > 1:
                self._desenhar_rodape(indice, total)
            super().showPage()
        super().save()

    def _desenhar_rodape(self, indice: int, total: int) -> None:
        """Desenha a régua e os textos do rodapé institucional.

        Args:
            indice (int): Número da página corrente.
            total (int): Total de páginas do documento.
        """
        largura, _ = A4
        y = MARGEM_BASE - 8 * mm
        self.saveState()
        self.setStrokeColor(cor("linha"))
        self.setLineWidth(0.25)
        self.line(MARGEM_LATERAL, y + 4 * mm, largura - MARGEM_LATERAL, y + 4 * mm)
        self.setFont("Helvetica", 7)
        self.setFillColor(cor("tinta_sec"))
        self.drawString(MARGEM_LATERAL, y, RODAPE_INSTITUCIONAL)
        self.drawRightString(largura - MARGEM_LATERAL, y, f"Página {indice} de {total}")
        self.restoreState()


# =============================================================================
# PDF — BLOCOS DE CONTEÚDO
# =============================================================================


def _bloco_sumario(escopo: str, largura: float, estilos: dict) -> list:
    """Monta o sumário de conteúdo exibido na capa.

    Args:
        escopo (str): Escopo do relatório, que define quais seções existem.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables do sumário.
    """
    secoes = []
    if escopo in ("geral", "completo"):
        secoes += [
            ("Panorama do período", "Receitas x despesas mês a mês em gráfico"),
            ("Resultado mês a mês", "Consolidação mensal de receitas, despesas e resultado"),
            ("Despesas por categoria", "Categorias com maior peso no período"),
        ]
    if escopo in ("investimentos", "completo"):
        secoes += [
            ("Alocação da carteira", "Distribuição do patrimônio por classe de ativo"),
            ("Carteira de investimentos", "Posições, metas de alocação e sugestão de aporte"),
            ("Proventos recebidos", "Dividendos e juros sobre capital por ativo"),
        ]
    if escopo in ("geral", "completo"):
        secoes.append(("Anexo A · Extrato detalhado", "Todos os lançamentos do período"))
    if escopo in ("investimentos", "completo"):
        secoes.append(
            ("Anexo B · Transações de investimento", "Ordens de compra, venda e proventos")
        )

    estilo_secao = _estilo_celula(8, "tinta", negrito=True)
    estilo_descricao = _estilo_celula(8, "tinta_sec")
    linhas = [
        [Paragraph(escape_xml(nome), estilo_secao), Paragraph(escape_xml(detalhe), estilo_descricao)]
        for nome, detalhe in secoes
    ]
    tabela = Table(linhas, colWidths=_larguras(largura, [0.36, 0.64]), hAlign="LEFT")
    tabela.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, -1), 0),
                ("LEFTPADDING", (1, 0), (1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LINEBELOW", (0, 0), (-1, -2), 0.25, cor("linha")),
            ]
        )
    )
    return [
        _texto("CONTEÚDO DESTE RELATÓRIO", estilos["kpi_rotulo_esq"]),
        Spacer(1, 2 * mm),
        tabela,
    ]


def _bloco_capa(usuario, data_inicio, data_fim, escopo, contexto, largura, estilos) -> list:
    """Monta a capa com título, identificação, indicadores e gráfico em cascata.

    Args:
        usuario (User): Usuário solicitante do relatório.
        data_inicio (date): Início do período.
        data_fim (date): Fim do período.
        escopo (str): Escopo do relatório.
        contexto (dict): Totais pré-calculados do período.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables da capa.
    """
    titular = (usuario.get_full_name() or usuario.username or "").strip()
    rotulo_escopo = {
        "geral": "Movimentações gerais",
        "investimentos": "Carteira de investimentos",
        "completo": "Movimentações gerais e carteira de investimentos",
    }.get(escopo, "Movimentações gerais e carteira de investimentos")

    elementos = [
        Spacer(1, 10 * mm),
        _texto("Relatório Financeiro", estilos["capa_titulo"]),
        _texto(
            f"Período de {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
            estilos["capa_periodo"],
        ),
        # Markup intencional: só a parte dinâmica é escapada.
        Paragraph(f"<b>Titular:</b> {escape_xml(titular) or '—'}", estilos["capa_meta"]),
        Paragraph(f"<b>Conteúdo:</b> {rotulo_escopo}", estilos["capa_meta"]),
        Paragraph(f"<b>Emissão:</b> {contexto['emissao']}", estilos["capa_meta"]),
        Spacer(1, 9 * mm),
    ]

    if escopo == "investimentos":
        indicadores = [
            ("Patrimônio investido", formatar_moeda(contexto["patrimonio"]), "marca"),
            ("Ativos em carteira", formatar_numero(contexto["qtd_ativos"], 0), "marca_esc"),
            (
                "Proventos no período",
                formatar_moeda(contexto["proventos"]),
                "positivo" if _dec(contexto["proventos"]) else "tinta",
            ),
            (
                "Resultado da carteira",
                formatar_moeda(contexto["resultado_carteira"]),
                "positivo"
                if _dec(contexto["resultado_carteira"]) > 0
                else ("negativo" if _dec(contexto["resultado_carteira"]) < 0 else "tinta"),
            ),
        ]
    else:
        indicadores = [
            (
                "Receitas do período",
                formatar_moeda(contexto["receitas"]),
                "positivo" if _dec(contexto["receitas"]) else "tinta",
            ),
            (
                "Despesas do período",
                formatar_moeda(contexto["despesas"]),
                "negativo" if _dec(contexto["despesas"]) else "tinta",
            ),
            (
                "Resultado do período",
                formatar_moeda(contexto["resultado"]),
                "positivo"
                if _dec(contexto["resultado"]) > 0
                else ("negativo" if _dec(contexto["resultado"]) < 0 else "tinta"),
            ),
            (
                "Patrimônio investido" if escopo == "completo" else "Lançamentos",
                formatar_moeda(contexto["patrimonio"])
                if escopo == "completo"
                else formatar_numero(contexto["qtd_lancamentos"], 0),
                "marca",
            ),
        ]

    elementos.append(_cartoes_kpi(indicadores, largura, estilos))

    if escopo in ("geral", "completo") and (
        _dec(contexto["receitas"]) or _dec(contexto["despesas"])
    ):
        elementos.append(Spacer(1, 10 * mm))
        elementos.append(
            render_grafico_waterfall(contexto["receitas"], contexto["despesas"], largura)
        )

    elementos.append(Spacer(1, 5 * mm))
    if escopo in ("geral", "completo"):
        elementos.append(
            _texto(
                "Os valores consolidam lançamentos de caixa e faturas de cartão de crédito, "
                "sem duplicar as despesas individuais que compõem cada fatura.",
                estilos["nota"],
            )
        )
    else:
        elementos.append(
            _texto(
                "Patrimônio e resultado refletem a posição da custódia na data de emissão, "
                "avaliada pela cotação mais recente de cada ativo.",
                estilos["nota"],
            )
        )
    elementos.append(Spacer(1, 11 * mm))
    elementos.extend(_bloco_sumario(escopo, largura, estilos))
    return elementos


def _bloco_graficos(comparativo, alocacao, largura, estilos) -> list:
    """Monta a página de gráficos (evolução mensal e alocação da carteira).

    Args:
        comparativo (list[dict]): Série mensal de receitas/despesas.
        alocacao (list[dict]): Distribuição da carteira por classe.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables da seção de gráficos, ou lista vazia se não houver dados.
    """
    elementos = []
    grafico_evolucao = render_grafico_evolucao_mensal(comparativo, largura)
    grafico_alocacao = render_grafico_alocacao(alocacao, largura)

    if grafico_evolucao is None and grafico_alocacao is None:
        return elementos

    titulo = "Panorama do período" if grafico_evolucao is not None else "Panorama da carteira"
    elementos.extend(_titulo_secao(titulo, largura, estilos))

    if grafico_evolucao is not None:
        elementos.append(grafico_evolucao)
        if len(comparativo) > MESES_NO_GRAFICO:
            elementos.append(
                _texto(
                    f"Gráfico limitado aos {MESES_NO_GRAFICO} meses mais recentes do período; "
                    "a série completa consta em “Resultado mês a mês”.",
                    estilos["nota"],
                )
            )
        elementos.append(Spacer(1, 8 * mm))

    if grafico_alocacao is not None:
        elementos.append(
            Paragraph("<b>Alocação da carteira por classe de ativo</b>", estilos["corpo"])
        )
        elementos.append(
            _texto("Posição da custódia na data de emissão do relatório.", estilos["nota"])
        )
        elementos.append(Spacer(1, 2 * mm))
        elementos.append(grafico_alocacao)

    return elementos


def _bloco_analise(comparativo, categorias, largura, estilos) -> list:
    """Monta as tabelas agregadas: resultado mês a mês e despesas por categoria.

    Args:
        comparativo (list[dict]): Série mensal de receitas/despesas.
        categorias (list[dict]): Despesas agrupadas por categoria.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables da seção de análise.
    """
    elementos = _titulo_secao("Resultado mês a mês", largura, estilos)

    if comparativo:
        linhas = [["PERÍODO", "RECEITAS", "DESPESAS", "RESULTADO"]]
        total_receitas = total_despesas = Decimal("0.00")
        for item in comparativo:
            total_receitas += _dec(item["receitas"])
            total_despesas += _dec(item["despesas"])
            linhas.append(
                [
                    formatar_mes_ano(item["periodo"]),
                    formatar_moeda(item["receitas"]),
                    formatar_moeda(item["despesas"]),
                    _valor_colorido(item["saldo"], fonte=8),
                ]
            )
        linhas.append(
            [
                "TOTAL DO PERÍODO",
                formatar_moeda(total_receitas),
                formatar_moeda(total_despesas),
                _valor_colorido(total_receitas - total_despesas, fonte=8, negrito=True),
            ]
        )
        elementos.append(
            _tabela(
                linhas,
                _larguras(largura, [0.28, 0.24, 0.24, 0.24]),
                numericas=(1, 2, 3),
                total=True,
                fonte=8,
            )
        )
    else:
        elementos.append(
            _texto("Nenhuma movimentação registrada no período.", estilos["vazio"])
        )

    elementos.append(Spacer(1, 9 * mm))
    elementos.extend(
        _titulo_secao("Despesas por categoria", largura, estilos)
    )

    if categorias:
        linhas = [["CATEGORIA", "TOTAL", "% DAS DESPESAS"]]
        total_categorias = Decimal("0.00")
        for item in categorias:
            total_categorias += _dec(item["total"])
            linhas.append(
                [
                    item["categoria"],
                    formatar_moeda(item["total"]),
                    formatar_percentual(item["percentual"]),
                ]
            )
        linhas.append(["TOTAL", formatar_moeda(total_categorias), formatar_percentual(100)])
        elementos.append(
            _tabela(
                linhas,
                _larguras(largura, [0.48, 0.28, 0.24]),
                numericas=(1, 2),
                total=True,
                fonte=8,
            )
        )
    else:
        elementos.append(_texto("Nenhuma despesa registrada no período.", estilos["vazio"]))

    return elementos


def _bloco_carteira(investimentos, largura, estilos) -> list:
    """Monta a tabela de posições da carteira com metas e sugestão de aporte.

    Args:
        investimentos (QuerySet): Ativos do usuário no período.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables da seção da carteira.
    """
    elementos = _titulo_secao(
        "Carteira de investimentos",
        largura,
        estilos,
        nota=(
            "Posições na data de emissão do relatório. Sugestão positiva indica quanto "
            "falta aportar para atingir a meta de alocação."
        ),
    )

    ativos = list(investimentos)
    if not ativos:
        elementos.append(_texto("Nenhum ativo em carteira no período.", estilos["vazio"]))
        return elementos

    total_mercado = sum((ativo.valor_total_atual for ativo in ativos), Decimal("0.00"))
    linhas = [
        [
            "TICKER", "NOME", "CLASSE", "QTD.", "P. MÉDIO", "MERCADO",
            "META", "IDEAL", "SUGESTÃO", "RESULTADO",
        ]
    ]
    total_investido = Decimal("0.00")
    total_ideal = Decimal("0.00")

    for ativo in ativos:
        valor_mercado = ativo.valor_total_atual
        valor_investido = ativo.valor_investido
        meta = _dec(ativo.meta_porcentagem)
        valor_ideal = (meta / 100) * total_mercado if total_mercado > 0 else Decimal("0.00")
        total_investido += valor_investido
        total_ideal += valor_ideal
        classe = (
            ativo.subcategoria.categoria.classe.nome if ativo.subcategoria else "Sem classe"
        )
        linhas.append(
            [
                ativo.ticker,
                ativo.nome or "—",
                classe,
                formatar_numero(ativo.quantidade),
                formatar_moeda(ativo.preco_medio),
                formatar_moeda(valor_mercado),
                formatar_percentual(meta),
                formatar_moeda(valor_ideal),
                _valor_colorido(valor_ideal - valor_mercado, fonte=6.5),
                _valor_colorido(valor_mercado - valor_investido, fonte=6.5),
            ]
        )

    linhas.append(
        [
            "TOTAL", "", "", "", "",
            formatar_moeda(total_mercado),
            "",
            formatar_moeda(total_ideal),
            formatar_moeda(total_ideal - total_mercado),
            _valor_colorido(total_mercado - total_investido, fonte=6.5, negrito=True),
        ]
    )

    elementos.append(
        _tabela(
            linhas,
            _larguras(
                largura,
                [0.135, 0.125, 0.12, 0.075, 0.085, 0.095, 0.07, 0.095, 0.095, 0.095],
            ),
            numericas=(3, 4, 5, 6, 7, 8, 9),
            total=True,
            fonte=6.5,
        )
    )
    return elementos


def _bloco_proventos(proventos, largura, estilos) -> list:
    """Monta a tabela de proventos recebidos por ativo no período.

    Args:
        proventos (Iterable[dict]): Saída de `get_proventos_data`.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables da seção de proventos.
    """
    elementos = _titulo_secao("Proventos recebidos", largura, estilos)
    itens = list(proventos)

    if not itens:
        elementos.append(
            _texto("Nenhum provento recebido no período.", estilos["vazio"])
        )
        return elementos

    linhas = [["TICKER", "TOTAL RECEBIDO"]]
    total = Decimal("0.00")
    for item in itens:
        total += _dec(item["total"])
        linhas.append([item["ativo__ticker"], formatar_moeda(item["total"])])
    linhas.append(["TOTAL", formatar_moeda(total)])

    elementos.append(
        _tabela(
            linhas,
            _larguras(largura, [0.5, 0.5]),
            numericas=(1,),
            total=True,
            fonte=8,
        )
    )
    return elementos


def _anexo_movimentacoes(movimentacoes, largura, estilos) -> list:
    """Monta o anexo com o extrato detalhado de todos os lançamentos do período.

    Despesas aparecem com sinal negativo para que a última linha totalize o
    resultado líquido do período.

    Args:
        movimentacoes (QuerySet): Lançamentos do período.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables do anexo.
    """
    elementos = _titulo_secao(
        "Anexo A · Extrato detalhado",
        largura,
        estilos,
        nota="Despesas são exibidas com sinal negativo; o total corresponde ao resultado líquido.",
    )

    linhas = [["DATA", "TIPO", "DESCRIÇÃO", "CATEGORIA", "VALOR", "SITUAÇÃO"]]
    liquido = Decimal("0.00")
    quantidade = 0

    for movimentacao in movimentacoes:
        quantidade += 1
        receita = movimentacao.tipo == Conta.TIPO_RECEITA
        valor = _dec(movimentacao.valor) if receita else -_dec(movimentacao.valor)
        liquido += valor
        linhas.append(
            [
                movimentacao.data_prevista.strftime("%d/%m/%Y"),
                "Receita" if receita else "Despesa",
                movimentacao.descricao,
                movimentacao.categoria.nome if movimentacao.categoria else "Sem categoria",
                formatar_moeda(valor),
                "Realizada" if movimentacao.transacao_realizada else "Pendente",
            ]
        )

    if quantidade == 0:
        elementos.append(
            _texto("Nenhuma movimentação registrada no período.", estilos["vazio"])
        )
        return elementos

    linhas.append(
        [f"{formatar_numero(quantidade, 0)} lançamentos", "", "", "", formatar_moeda(liquido), ""]
    )
    elementos.append(
        _tabela(
            linhas,
            _larguras(largura, [0.11, 0.09, 0.34, 0.19, 0.15, 0.12]),
            numericas=(4,),
            total=True,
            fonte=7.5,
        )
    )
    return elementos


def _anexo_transacoes(transacoes, largura, estilos) -> list:
    """Monta o anexo com o histórico de ordens de investimento do período.

    Args:
        transacoes (QuerySet): Ordens executadas no período.
        largura (float): Largura útil do frame.
        estilos (dict): Mapa de estilos retornado por `_estilos`.

    Returns:
        list: Flowables do anexo.
    """
    elementos = _titulo_secao("Anexo B · Histórico de transações de investimento", largura, estilos)

    linhas = [["DATA", "TICKER", "OPERAÇÃO", "QTD.", "PREÇO", "TAXAS", "TOTAL"]]
    total = Decimal("0.00")
    quantidade = 0

    for transacao in transacoes:
        quantidade += 1
        total += _dec(transacao.valor_total)
        linhas.append(
            [
                transacao.data.strftime("%d/%m/%Y"),
                transacao.ativo.ticker,
                transacao.get_tipo_display(),
                formatar_numero(transacao.quantidade),
                formatar_moeda(transacao.preco_unitario),
                formatar_moeda(transacao.taxas),
                formatar_moeda(transacao.valor_total),
            ]
        )

    if quantidade == 0:
        elementos.append(
            _texto("Nenhuma transação de investimento no período.", estilos["vazio"])
        )
        return elementos

    linhas.append(
        [f"{formatar_numero(quantidade, 0)} ordens", "", "", "", "", "", formatar_moeda(total)]
    )
    elementos.append(
        _tabela(
            linhas,
            _larguras(largura, [0.13, 0.17, 0.16, 0.12, 0.14, 0.13, 0.15]),
            numericas=(3, 4, 5, 6),
            total=True,
            fonte=7.5,
        )
    )
    return elementos


def gerar_pdf(usuario, data_inicio: date, data_fim: date, escopo: str = "completo") -> bytes:
    """Gera o relatório financeiro em PDF com capa, resumo executivo, gráficos e anexos.

    A estrutura vai do agregado ao detalhe: capa com indicadores e composição do
    resultado, panorama gráfico do período, tabelas de análise, carteira e
    proventos, e por fim os anexos com o extrato completo. Todas as páginas
    internas trazem cabeçalho da marca e rodapé com "Página X de Y", e os
    cabeçalhos de tabela se repetem nas quebras de página.

    Args:
        usuario (User): Instância do usuário Django solicitante.
        data_inicio (date): Limite de início para filtragem do relatório.
        data_fim (date): Limite final para filtragem do relatório.
        escopo (str, optional): Escopo do relatório ('geral', 'investimentos', 'completo'). Defaults to "completo".

    Returns:
        bytes: O conteúdo em bytes do arquivo PDF gerado.
    """
    if escopo not in ("geral", "investimentos", "completo"):
        escopo = "completo"

    inclui_geral = escopo in ("geral", "completo")
    inclui_investimentos = escopo in ("investimentos", "completo")

    periodo_texto = (
        f"Relatório Financeiro · {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    )
    emissao = timezone.localtime().strftime("%d/%m/%Y às %H:%M")

    comparativo = get_comparativo_mensal_data(usuario, data_inicio, data_fim) if inclui_geral else []
    categorias = get_despesas_por_categoria(usuario, data_inicio, data_fim) if inclui_geral else []
    movimentacoes = get_movimentacoes(usuario, data_inicio, data_fim) if inclui_geral else []

    investimentos = (
        list(get_investimentos(usuario, data_inicio, data_fim)) if inclui_investimentos else []
    )
    transacoes_invest = (
        get_transacoes_investimento(usuario, data_inicio, data_fim) if inclui_investimentos else []
    )
    proventos = (
        list(get_proventos_data(usuario, data_inicio, data_fim)) if inclui_investimentos else []
    )
    alocacao = get_alocacao_data(usuario, data_fim) if inclui_investimentos else []

    total_receitas = sum((_dec(item["receitas"]) for item in comparativo), Decimal("0.00"))
    total_despesas = sum((_dec(item["despesas"]) for item in comparativo), Decimal("0.00"))
    patrimonio = sum((_dec(item["valor"]) for item in alocacao), Decimal("0.00"))
    contexto = {
        "emissao": emissao,
        "receitas": total_receitas,
        "despesas": total_despesas,
        "resultado": total_receitas - total_despesas,
        "patrimonio": patrimonio,
        "qtd_ativos": len(investimentos),
        "qtd_lancamentos": movimentacoes.count() if inclui_geral else 0,
        "proventos": sum((_dec(item["total"]) for item in proventos), Decimal("0.00")),
        "resultado_carteira": sum(
            (ativo.valor_total_atual - ativo.valor_investido for ativo in investimentos),
            Decimal("0.00"),
        ),
    }

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=MARGEM_LATERAL,
        rightMargin=MARGEM_LATERAL,
        topMargin=MARGEM_TOPO,
        bottomMargin=MARGEM_BASE,
        title=f"Relatório Financeiro {data_inicio.strftime('%d-%m-%Y')} a {data_fim.strftime('%d-%m-%Y')}",
        author="FreeCash",
        subject=f"Relatório financeiro do período de {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
        creator="FreeCash",
    )

    estilos = _estilos()
    largura = doc.width
    elementos = _bloco_capa(
        usuario, data_inicio, data_fim, escopo, contexto, largura, estilos
    )

    graficos = _bloco_graficos(comparativo, alocacao, largura, estilos)
    if graficos:
        elementos.append(PageBreak())
        elementos.extend(graficos)

    if inclui_geral:
        # Sem quebra forçada: a análise ocupa o espaço restante da página de gráficos.
        elementos.append(Spacer(1, 10 * mm) if graficos else PageBreak())
        elementos.extend(_bloco_analise(comparativo, categorias, largura, estilos))

    if inclui_investimentos:
        # Só quebra página depois das tabelas longas da análise geral; no escopo de
        # investimentos a carteira aproveita o espaço restante da página do gráfico.
        elementos.append(PageBreak() if inclui_geral or not graficos else Spacer(1, 9 * mm))
        elementos.extend(_bloco_carteira(investimentos, largura, estilos))
        elementos.append(Spacer(1, 9 * mm))
        elementos.extend(_bloco_proventos(proventos, largura, estilos))

    if inclui_geral:
        elementos.append(PageBreak())
        elementos.extend(_anexo_movimentacoes(movimentacoes, largura, estilos))

    if inclui_investimentos:
        elementos.append(PageBreak())
        elementos.extend(_anexo_transacoes(transacoes_invest, largura, estilos))

    capa_callback, interna_callback = _construir_moldura(periodo_texto, f"Emitido em {emissao}")
    doc.build(
        elementos,
        onFirstPage=capa_callback,
        onLaterPages=interna_callback,
        canvasmaker=CanvasNumerado,
    )
    buffer.seek(0)
    return buffer.getvalue()
