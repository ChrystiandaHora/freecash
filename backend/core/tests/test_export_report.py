"""Testes dos relatórios exportáveis (PDF, Excel e CSV) e da camada de formatação.

Cobre quatro frentes que os geradores anteriores não tinham:

1. Formatação pt-BR — a falha mais visível do relatório antigo era imprimir
   ``R$ 3,048.21`` (padrão americano) em um documento em português.
2. `get_despesas_por_categoria` — nova agregação, que precisa respeitar o mesmo
   filtro anti-duplicidade de fatura de cartão usado pelas movimentações.
3. `gerar_pdf` — geração ponta a ponta nos três escopos, incluindo os casos de
   base vazia, documento multipágina e isolamento entre usuários.
4. `gerar_excel` / `gerar_csv` — números gravados como número (e não texto),
   totais que fecham com os dados e CSV que o Excel pt-BR abre sem assistente.
"""

import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import load_workbook

from core.models import CartaoCredito, Categoria, Conta
from core.services.export_report_service import (
    FORMATO_MOEDA,
    formatar_mes_ano,
    formatar_moeda,
    formatar_numero,
    formatar_percentual,
    gerar_csv,
    gerar_excel,
    gerar_pdf,
    get_despesas_por_categoria,
)
from investimento.models import (
    Ativo,
    ClasseAtivo,
    CategoriaAtivo,
    SubcategoriaAtivo,
    Transacao,
)

INICIO = date(2026, 1, 1)
FIM = date(2026, 12, 31)


class FormatacaoPtBrTests(TestCase):
    """Formatação numérica e de datas no padrão brasileiro."""

    def test_moeda_usa_ponto_de_milhar_e_virgula_decimal(self):
        self.assertEqual(formatar_moeda(Decimal("3048.21")), "R$ 3.048,21")
        self.assertEqual(formatar_moeda(Decimal("1234567.89")), "R$ 1.234.567,89")
        self.assertEqual(formatar_moeda(Decimal("0.5")), "R$ 0,50")

    def test_moeda_negativa_tem_sinal_antes_do_simbolo(self):
        self.assertEqual(formatar_moeda(Decimal("-250")), "-R$ 250,00")

    def test_moeda_aceita_none_zero_e_float(self):
        self.assertEqual(formatar_moeda(None), "R$ 0,00")
        self.assertEqual(formatar_moeda(0), "R$ 0,00")
        self.assertEqual(formatar_moeda(1234.5), "R$ 1.234,50")

    def test_moeda_sem_simbolo(self):
        self.assertEqual(formatar_moeda(Decimal("99.9"), simbolo=False), "99,90")

    def test_numero_e_percentual(self):
        self.assertEqual(formatar_numero(Decimal("1234.5"), 0), "1.234")
        self.assertEqual(formatar_numero(Decimal("317.82")), "317,82")
        self.assertEqual(formatar_percentual(Decimal("70.14")), "70,1%")

    def test_mes_ano_legivel(self):
        self.assertEqual(formatar_mes_ano("2026-01"), "jan/2026")
        self.assertEqual(formatar_mes_ano("2025-12"), "dez/2025")

    def test_mes_ano_devolve_entrada_quando_invalida(self):
        self.assertEqual(formatar_mes_ano("indefinido"), "indefinido")
        self.assertEqual(formatar_mes_ano(None), "None")


class BaseRelatorioTestCase(TestCase):
    """Cenário compartilhado: um usuário com movimentações e carteira."""

    def setUp(self):
        self.user = User.objects.create_user(username="chrystian", password="senha-forte-123")
        self.outro = User.objects.create_user(username="terceiro", password="senha-forte-123")

        self.moradia = Categoria.objects.create(
            usuario=self.user, nome="Moradia", tipo=Conta.TIPO_DESPESA
        )
        self.lazer = Categoria.objects.create(
            usuario=self.user, nome="Lazer", tipo=Conta.TIPO_DESPESA
        )
        self.salario = Categoria.objects.create(
            usuario=self.user, nome="Salário", tipo=Conta.TIPO_RECEITA
        )
        self.cartao = CartaoCredito.objects.create(usuario=self.user, nome="Cartão Teste")

        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_RECEITA, descricao="Salário",
            valor=Decimal("8093.68"), data_prevista=date(2026, 3, 5),
            categoria=self.salario, transacao_realizada=True,
        )
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Condomínio",
            valor=Decimal("430.00"), data_prevista=date(2026, 3, 10),
            categoria=self.moradia, transacao_realizada=True,
        )
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Aluguel",
            valor=Decimal("1200.00"), data_prevista=date(2026, 4, 10),
            categoria=self.moradia,
        )
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Cinema",
            valor=Decimal("70.00"), data_prevista=date(2026, 4, 15),
            categoria=self.lazer,
        )
        # Fatura consolidada: entra no relatório
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Fatura Cartão Teste - 04/2026",
            valor=Decimal("500.00"), data_prevista=date(2026, 4, 10),
            cartao=self.cartao, eh_fatura_cartao=True, categoria=self.lazer,
        )
        # Despesa individual do cartão: NÃO entra (já está dentro da fatura acima)
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Compra no cartão",
            valor=Decimal("500.00"), data_prevista=date(2026, 4, 10),
            cartao=self.cartao, categoria=self.lazer,
        )
        # Fora do período consultado
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Despesa antiga",
            valor=Decimal("999.00"), data_prevista=date(2025, 5, 10),
            categoria=self.moradia,
        )
        # De outro usuário: nunca deve aparecer
        Conta.objects.create(
            usuario=self.outro, tipo=Conta.TIPO_DESPESA, descricao="Despesa de terceiro",
            valor=Decimal("4242.00"), data_prevista=date(2026, 3, 10),
        )

    def criar_carteira(self):
        """Cria um ativo com transações para exercitar o escopo de investimentos."""
        # A criação do usuário já semeia a árvore ANBIMA padrão; reaproveita se existir.
        classe, _ = ClasseAtivo.objects.get_or_create(
            usuario=self.user, nome="Renda Variável"
        )
        categoria, _ = CategoriaAtivo.objects.get_or_create(
            usuario=self.user, classe=classe, nome="FIIs"
        )
        subcategoria, _ = SubcategoriaAtivo.objects.get_or_create(
            usuario=self.user, categoria=categoria, nome="Tijolo"
        )
        ativo = Ativo.objects.create(
            usuario=self.user,
            ticker="XPML11",
            nome="XP Malls FII",
            subcategoria=subcategoria,
            quantidade=Decimal("6"),
            preco_medio=Decimal("106.60"),
            meta_porcentagem=Decimal("5"),
        )
        Transacao.objects.create(
            usuario=self.user, ativo=ativo, tipo=Transacao.TIPO_COMPRA,
            data=date(2026, 2, 13), quantidade=Decimal("1"),
            preco_unitario=Decimal("110.86"), valor_total=Decimal("110.86"),
        )
        Transacao.objects.create(
            usuario=self.user, ativo=ativo, tipo=Transacao.TIPO_DIVIDENDO,
            data=date(2026, 3, 24), quantidade=Decimal("1"),
            preco_unitario=Decimal("4.60"), valor_total=Decimal("4.60"),
        )
        return ativo


class DespesasPorCategoriaTests(BaseRelatorioTestCase):
    """Agregação de despesas por categoria."""

    def test_agrupa_e_ordena_do_maior_para_o_menor(self):
        resultado = get_despesas_por_categoria(self.user, INICIO, FIM)
        nomes = [item["categoria"] for item in resultado]

        self.assertEqual(nomes, ["Moradia", "Lazer"])
        # 430 + 1200; a despesa de 999 está fora do período
        self.assertEqual(resultado[0]["total"], Decimal("1630.00"))
        # 70 (cinema) + 500 (fatura), sem a despesa individual do cartão
        self.assertEqual(resultado[1]["total"], Decimal("570.00"))

    def test_percentuais_somam_cem(self):
        resultado = get_despesas_por_categoria(self.user, INICIO, FIM)
        total = sum(item["percentual"] for item in resultado)
        self.assertAlmostEqual(float(total), 100.0, places=2)

    def test_despesa_individual_de_cartao_nao_duplica_a_fatura(self):
        total = sum(
            item["total"] for item in get_despesas_por_categoria(self.user, INICIO, FIM)
        )
        # 430 + 1200 + 70 + 500 (fatura). A "Compra no cartão" de 500 é ignorada.
        self.assertEqual(total, Decimal("2200.00"))

    def test_categoria_nula_recebe_rotulo_legivel(self):
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao="Sem classificação",
            valor=Decimal("10.00"), data_prevista=date(2026, 5, 1),
        )
        nomes = [item["categoria"] for item in get_despesas_por_categoria(self.user, INICIO, FIM)]
        self.assertIn("Sem categoria", nomes)

    def test_agrupa_excedente_em_outras_categorias(self):
        for indice in range(6):
            categoria = Categoria.objects.create(
                usuario=self.user, nome=f"Categoria {indice}", tipo=Conta.TIPO_DESPESA
            )
            Conta.objects.create(
                usuario=self.user, tipo=Conta.TIPO_DESPESA, descricao=f"Gasto {indice}",
                valor=Decimal("5.00"), data_prevista=date(2026, 6, 1), categoria=categoria,
            )

        resultado = get_despesas_por_categoria(self.user, INICIO, FIM, limite=3)

        self.assertEqual(len(resultado), 4)
        self.assertEqual(resultado[-1]["categoria"], "Outras categorias")
        # Nada se perde no agrupamento
        self.assertEqual(
            sum(item["total"] for item in resultado), Decimal("2230.00")
        )

    def test_nao_enxerga_dados_de_outro_usuario(self):
        resultado = get_despesas_por_categoria(self.outro, INICIO, FIM)
        self.assertEqual(len(resultado), 1)
        self.assertEqual(resultado[0]["total"], Decimal("4242.00"))


class GerarPdfTests(BaseRelatorioTestCase):
    """Geração do documento PDF ponta a ponta."""

    def _paginas(self, payload: bytes) -> int:
        """Conta as páginas declaradas no PDF.

        Args:
            payload (bytes): Conteúdo do arquivo gerado.

        Returns:
            int: Quantidade de objetos de página encontrados.
        """
        return payload.count(b"/Type /Page\n") + payload.count(b"/Type /Page ")

    def test_gera_pdf_valido_nos_tres_escopos(self):
        self.criar_carteira()
        for escopo in ("geral", "investimentos", "completo"):
            with self.subTest(escopo=escopo):
                payload = gerar_pdf(self.user, INICIO, FIM, escopo)
                self.assertIsInstance(payload, bytes)
                self.assertTrue(payload.startswith(b"%PDF"))
                self.assertTrue(payload.rstrip().endswith(b"%%EOF"))
                self.assertGreaterEqual(self._paginas(payload), 1)

    def test_escopo_invalido_cai_para_completo(self):
        self.criar_carteira()
        self.assertEqual(
            len(gerar_pdf(self.user, INICIO, FIM, "escopo-inexistente")),
            len(gerar_pdf(self.user, INICIO, FIM, "completo")),
        )

    def test_usuario_sem_dados_nao_quebra(self):
        vazio = User.objects.create_user(username="novato", password="senha-forte-123")
        payload = gerar_pdf(vazio, INICIO, FIM, "completo")
        self.assertTrue(payload.startswith(b"%PDF"))

    def test_periodo_sem_movimentacoes_nao_quebra(self):
        payload = gerar_pdf(self.user, date(2019, 1, 1), date(2019, 12, 31), "completo")
        self.assertTrue(payload.startswith(b"%PDF"))

    def test_volume_alto_gera_documento_multipagina(self):
        for dia in range(1, 29):
            for mes in (7, 8, 9):
                Conta.objects.create(
                    usuario=self.user, tipo=Conta.TIPO_DESPESA,
                    descricao=f"Lançamento {mes}-{dia}", valor=Decimal("12.34"),
                    data_prevista=date(2026, mes, dia), categoria=self.lazer,
                )

        payload = gerar_pdf(self.user, INICIO, FIM, "geral")
        self.assertGreater(self._paginas(payload), 3)

    def test_descricao_com_caracteres_de_marcacao_nao_quebra_o_documento(self):
        Conta.objects.create(
            usuario=self.user, tipo=Conta.TIPO_DESPESA,
            descricao="Consultoria <b>Alfa</b> & Beta <não fecha",
            valor=Decimal("15.00"), data_prevista=date(2026, 5, 20), categoria=self.lazer,
        )
        payload = gerar_pdf(self.user, INICIO, FIM, "geral")
        self.assertTrue(payload.startswith(b"%PDF"))

    def test_titulo_do_documento_traz_o_periodo(self):
        payload = gerar_pdf(self.user, INICIO, FIM, "geral")
        self.assertIn(b"01-01-2026", payload)


class GerarExcelTests(BaseRelatorioTestCase):
    """Planilha Excel: estrutura das abas, tipos nativos e totais."""

    def _abrir(self, escopo="completo"):
        """Gera a planilha e devolve a pasta de trabalho carregada.

        Args:
            escopo (str, optional): Escopo do relatório. Defaults to "completo".

        Returns:
            Workbook: Planilha gerada, pronta para inspeção.
        """
        return load_workbook(io.BytesIO(gerar_excel(self.user, INICIO, FIM, escopo)))

    def test_abas_por_escopo(self):
        self.criar_carteira()
        self.assertEqual(
            self._abrir("geral").sheetnames,
            ["Resumo", "Movimentações", "Resumo Mensal", "Despesas por Categoria"],
        )
        self.assertEqual(
            self._abrir("investimentos").sheetnames,
            ["Resumo", "Carteira", "Alocação", "Proventos", "Transações Invest."],
        )
        self.assertEqual(len(self._abrir("completo").sheetnames), 8)

    def test_resumo_e_a_primeira_aba_com_indicadores_como_numero(self):
        ws = self._abrir("geral")["Resumo"]
        rotulos = {
            linha[0].value: linha[1]
            for linha in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2)
        }
        receitas = rotulos["Receitas do período"]
        self.assertIsInstance(receitas.value, (int, float))
        self.assertAlmostEqual(receitas.value, 8093.68, places=2)
        self.assertEqual(receitas.number_format, FORMATO_MOEDA)
        # Contagem não deve ser formatada como moeda
        self.assertEqual(rotulos["Lançamentos no período"].number_format, "#,##0")

    def test_movimentacoes_guarda_data_e_valor_nativos(self):
        ws = self._abrir("geral")["Movimentações"]
        primeira = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(primeira[0].date(), date(2026, 3, 5))
        self.assertIsInstance(primeira[4], (int, float))
        self.assertEqual(ws.cell(row=2, column=1).number_format, "dd/mm/yyyy")
        self.assertEqual(ws.cell(row=2, column=5).number_format, FORMATO_MOEDA)

    def test_despesas_entram_negativas_e_o_total_fecha_o_liquido(self):
        ws = self._abrir("geral")["Movimentações"]
        valores = [
            linha[4] for linha in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True)
        ]
        self.assertTrue(any(valor < 0 for valor in valores), "despesas devem ser negativas")
        # 8093,68 de receita menos 2200,00 de despesas consideradas
        self.assertAlmostEqual(sum(valores), 5893.68, places=2)
        self.assertAlmostEqual(ws.cell(row=ws.max_row, column=5).value, 5893.68, places=2)

    def test_cabecalho_congelado_e_filtro_nao_alcancam_a_linha_de_total(self):
        ws = self._abrir("geral")["Movimentações"]
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertEqual(ws.auto_filter.ref, f"A1:F{ws.max_row - 1}")

    def test_linha_de_total_nao_recebe_formato_em_celula_de_texto(self):
        ws = self._abrir("geral")["Movimentações"]
        rotulo = ws.cell(row=ws.max_row, column=1)
        self.assertIn("lançamentos", str(rotulo.value))
        self.assertNotEqual(rotulo.number_format, "dd/mm/yyyy")

    def test_percentuais_sao_fracao_com_formato_percentual(self):
        ws = self._abrir("geral")["Despesas por Categoria"]
        celula = ws.cell(row=2, column=3)
        self.assertLessEqual(celula.value, 1.0)
        self.assertEqual(celula.number_format, "0.0%")

    def test_resumo_mensal_usa_data_real_para_ordenar_corretamente(self):
        ws = self._abrir("geral")["Resumo Mensal"]
        meses = [
            linha[0] for linha in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True)
        ]
        self.assertTrue(all(hasattr(mes, "year") for mes in meses))
        self.assertEqual(meses, sorted(meses))
        self.assertEqual(ws.cell(row=2, column=1).number_format, "mmm/yyyy")

    def test_carteira_totaliza_ideal_e_sugestao_somando_as_linhas(self):
        self.criar_carteira()
        ws = self._abrir("investimentos")["Carteira"]
        linhas = list(ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True))
        total = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
        for coluna in (9, 10):
            self.assertAlmostEqual(
                total[coluna], sum(linha[coluna] for linha in linhas), places=6
            )

    def test_usuario_sem_dados_gera_planilha_valida(self):
        vazio = User.objects.create_user(username="novato", password="senha-forte-123")
        wb = load_workbook(io.BytesIO(gerar_excel(vazio, INICIO, FIM, "completo")))
        self.assertIn("Resumo", wb.sheetnames)
        self.assertEqual(wb["Movimentações"].max_row, 1)  # só o cabeçalho


class GerarCsvTests(BaseRelatorioTestCase):
    """CSV: compatibilidade com Excel/LibreOffice em português."""

    def test_comeca_com_bom_para_o_excel_reconhecer_utf8(self):
        conteudo = gerar_csv(self.user, INICIO, FIM, "geral")
        self.assertTrue(conteudo.startswith("﻿"))
        self.assertIn("Movimentações", conteudo)
        self.assertIn("Descrição", conteudo)

    def test_usa_ponto_e_virgula_e_virgula_decimal(self):
        conteudo = gerar_csv(self.user, INICIO, FIM, "geral")
        self.assertIn("Data;Tipo;Descrição;Categoria;Valor (R$);Situação", conteudo)
        self.assertIn("8.093,68", conteudo)
        self.assertNotIn("8093.68", conteudo)

    def test_traz_as_secoes_do_escopo(self):
        self.criar_carteira()
        completo = gerar_csv(self.user, INICIO, FIM, "completo")
        for secao in (
            "MOVIMENTAÇÕES",
            "RESUMO MENSAL",
            "DESPESAS POR CATEGORIA",
            "CARTEIRA DE INVESTIMENTOS",
            "ALOCAÇÃO POR CLASSE",
            "PROVENTOS RECEBIDOS",
            "TRANSAÇÕES DE INVESTIMENTO",
        ):
            self.assertIn(secao, completo)

        geral = gerar_csv(self.user, INICIO, FIM, "geral")
        self.assertIn("MOVIMENTAÇÕES", geral)
        self.assertNotIn("CARTEIRA DE INVESTIMENTOS", geral)

    def test_despesa_negativa_e_total_liquido(self):
        conteudo = gerar_csv(self.user, INICIO, FIM, "geral")
        self.assertIn("-430,00", conteudo)
        self.assertIn("5.893,68", conteudo)

    def test_periodo_sem_dados_informa_ausencia(self):
        conteudo = gerar_csv(self.user, date(2019, 1, 1), date(2019, 12, 31), "geral")
        self.assertIn("Nenhuma movimentação no período", conteudo)

    def test_escopo_invalido_cai_para_completo(self):
        def sem_emissao(conteudo: str) -> list:
            """Remove a linha de emissão, que muda a cada minuto."""
            return [
                linha for linha in conteudo.splitlines() if not linha.startswith("Emitido em")
            ]

        self.assertEqual(
            sem_emissao(gerar_csv(self.user, INICIO, FIM, "escopo-inexistente")),
            sem_emissao(gerar_csv(self.user, INICIO, FIM, "completo")),
        )
